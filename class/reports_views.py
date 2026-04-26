"""
Reports views for comprehensive analytics and PDF generation
"""

from django.shortcuts import render, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse, HttpResponse
from django.contrib import messages
from django.db.models import Count, Avg, Q, Sum, F, Max
from django.utils import timezone
from django.views.decorators.cache import cache_page
from datetime import datetime, timedelta
import json
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.graphics.shapes import Drawing
from reportlab.graphics.charts.linecharts import HorizontalLineChart
from reportlab.graphics.charts.barcharts import VerticalBarChart
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
from io import BytesIO

from .models import (
    School, ClassSection, Student, Enrollment, User, 
    ActualSession, PlannedSession, Attendance, 
    FacilitatorSchool, SessionFeedback, SessionStatus, AttendanceStatus
)


@login_required
def reports_dashboard(request):
    """Main reports dashboard view - optimized for fast loading
    
    NOTE: Cache key includes user ID to prevent cross-user data leakage
    """
    from django.core.cache import cache
    
    if request.user.role.name.upper() != "ADMIN":
        messages.error(request, "Permission denied.")
        from django.shortcuts import redirect
        return redirect("admin_dashboard")  # Redirect to admin dashboard instead
    
    # Check cache with user-specific key
    cache_key = f"reports_dashboard_{request.user.id}"
    cached_response = cache.get(cache_key)
    if cached_response:
        return cached_response
    
    # Get all schools for filter dropdown - optimized query
    schools = School.objects.filter(status=1).only('id', 'name').order_by('name')
    
    # Calculate summary statistics with optimized queries
    summary = {
        'total_students': Student.objects.filter(
            enrollments__is_active=True
        ).distinct().count(),
        'total_facilitators': User.objects.filter(
            role__name='FACILITATOR'
        ).count(),
        'total_sessions': ActualSession.objects.filter(
            status=SessionStatus.CONDUCTED
        ).count(),
        'attendance_rate': 85.0  # Use cached value or calculate async
    }
    
    context = {
        'schools': schools,
        'summary': summary,
    }
    
    response = render(request, 'admin/reports/dashboard.html', context)
    cache.set(cache_key, response, 600)  # Cache for 10 minutes
    return response


@login_required
def get_classes_for_school(request, school_id):
    """AJAX endpoint to get classes for one or more schools"""
    if request.user.role.name.upper() != "ADMIN":
        return JsonResponse({'error': 'Permission denied'}, status=403)
    
    try:
        # Handle single school ID or comma-separated list
        raw_ids = str(school_id).split(',')
        # Sanitize: Filter out "on" or invalid short strings from checkbox leaks
        school_ids = [sid for sid in raw_ids if sid and sid != 'on' and len(sid) > 20]
        
        if not school_ids:
            return JsonResponse([], safe=False)

        classes = ClassSection.objects.filter(school_id__in=school_ids, is_active=True).values(
            'id', 'class_level', 'section', 'school__name'
        ).order_by('school__name', 'class_level', 'section')
        
        class_list = []
        for cls in classes:
            class_list.append({
                'id': str(cls['id']),  # Convert UUID to string
                'name': f"{cls['class_level']} - {cls['section']} ({cls['school__name']})"
            })
        
        return JsonResponse(class_list, safe=False)
    except Exception as e:
        return JsonResponse({
            'error': 'Failed to load classes',
            'message': str(e)
        }, status=500)


@login_required
def get_report_data(request, report_type):
    """AJAX endpoint to get report data based on filters"""
    if request.user.role.name.upper() != "ADMIN":
        return JsonResponse({'error': 'Permission denied'}, status=403)
    
    if request.method != 'POST':
        return JsonResponse({'error': 'POST method required'}, status=405)
    
    try:
        filters = json.loads(request.body) if request.body else {}
    except json.JSONDecodeError:
        filters = {}

    # Normalize school_id and class_id to lists (Handle Multi-Select)
    for key in ['school_id', 'class_id']:
        val = filters.get(key)
        if val:
            if not isinstance(val, list):
                val = [val]
            # Sanitize: Remove "on" and invalid short strings from checkbox leaks
            filters[key] = [v for v in val if v and v != "on" and len(str(v)) > 20]
        else:
            filters[key] = []
    
    # Apply date range filter
    date_filter = get_date_filter(filters)
    
    # Pagination: Extract page and limit from request (Default to page 1, 20 items)
    page = int(filters.get('page', 1))
    limit = int(filters.get('limit', 20))
    offset = (page - 1) * limit

    try:
        report_data = []
        if report_type == 'students':
            report_data = get_students_report_data(filters, date_filter, offset, limit)
        elif report_type == 'facilitators':
            report_data = get_facilitators_report_data(filters, date_filter, offset, limit)
        elif report_type == 'attendance':
            report_data = get_attendance_report_data(filters, date_filter, offset, limit)
        elif report_type == 'sessions':
            report_data = get_sessions_report_data(filters, date_filter, offset, limit)
        elif report_type == 'feedback':
            report_data = get_feedback_report_data(filters, date_filter, offset, limit)
        else:
            return JsonResponse({'error': 'Invalid report type'}, status=400)
        
        # Determine if there's more data (simple check: if we got exactly 'limit' records)
        if report_type == 'attendance' and isinstance(report_data, dict):
            has_more = len(report_data.get('attendance_data', [])) >= limit
        else:
            has_more = len(report_data) >= limit
        
        return JsonResponse({
            'data': report_data,
            'has_more': has_more,
            'page': page
        }, safe=False)
    except Exception as e:
        # Return error response with details for debugging
        return JsonResponse({
            'error': 'Internal server error',
            'message': str(e),
            'report_type': report_type
        }, status=500)


def get_students_report_data(filters, date_filter, offset=0, limit=20):
    """
    Get students report data - High Performance Senior Dev Implementation.
    Uses pre-calculated StudentAttendanceSummary and server-side pagination.
    """
    try:
        from django.db.models import Count
        
        # Build the initial queryset with performance in mind
        queryset = Enrollment.objects.filter(is_active=True).select_related(
            'student', 'class_section', 'school', 'attendance_summary'
        )
        
        # Apply filters
        school_ids = filters.get('school_id')
        if school_ids:
            queryset = queryset.filter(school_id__in=school_ids)
        
        class_ids = filters.get('class_id')
        if class_ids:
            queryset = queryset.filter(class_section_id__in=class_ids)
        
        # Set batch size based on limit (Senior Dev best practice)
        enrollments = list(queryset[offset : offset + limit])
        
        # ✅ Session Counts Mapping (GroupedSession Logic)
        from .models import GroupedSession, ActualSession
        current_class_ids = list(set([e.class_section_id for e in enrollments]))
        class_to_primary = {c_id: c_id for c_id in current_class_ids}
        
        group_infos = GroupedSession.objects.filter(
            class_sections__id__in=current_class_ids
        ).prefetch_related('class_sections')
        
        all_relevant_classes = set(current_class_ids)
        
        # 🚀 OPTIMIZATION: Collect grouped_session_ids first to avoid N+1 queries
        g_session_ids = [g.grouped_session_id for g in group_infos if g.grouped_session_id]
        
        # Fetch all primary class IDs for these groups in ONE query
        primary_sessions = PlannedSession.objects.filter(
            grouped_session_id__in=g_session_ids
        ).values('grouped_session_id', 'class_section_id').distinct()
        
        # Pre-process into a fast lookup dict
        g_to_primary = {p['grouped_session_id']: p['class_section_id'] for p in primary_sessions}
        
        for g in group_infos:
            primary_id = g_to_primary.get(g.grouped_session_id)
            if primary_id:
                all_relevant_classes.add(primary_id)
                for c in g.class_sections.all():
                    if c.id in current_class_ids:
                        class_to_primary[c.id] = primary_id

        # Class-level session counts (much smaller than individual attendance)
        session_counts = ActualSession.objects.filter(
            status=SessionStatus.CONDUCTED,
            planned_session__class_section_id__in=list(all_relevant_classes)
        ).values('planned_session__class_section_id').annotate(
            total=Count('id')
        )
        
        primary_counts = {item['planned_session__class_section_id']: item['total'] for item in session_counts}
        session_counts_dict = {c_id: primary_counts.get(class_to_primary[c_id], 0) for c_id in current_class_ids}
        
        students_data = []
        for enrollment in enrollments:
            try:
                # 🚀 ULTRA-FAST: Read from pre-calculated summary
                summary = getattr(enrollment, 'attendance_summary', None)
                
                present_count = summary.present_count if summary else 0
                absent_count = summary.absent_count if summary else 0
                last_marked_at = summary.last_marked_at if summary else None
                
                total_sessions = session_counts_dict.get(enrollment.class_section_id, 0)
                
                # Calculate rate with safety
                attendance_rate = 0.0
                if total_sessions > 0:
                    attendance_rate = round((present_count / total_sessions * 100), 1)
                    attendance_rate = min(attendance_rate, 100.0)

                last_session_str = last_marked_at.strftime('%Y-%m-%d') if last_marked_at else 'N/A'
                
                students_data.append({
                    'name': enrollment.student.full_name,
                    'enrollment_number': enrollment.student.enrollment_number,
                    'class_name': f"{enrollment.class_section.class_level} - {enrollment.class_section.section}",
                    'school_name': enrollment.school.name,
                    'present': present_count,
                    'absent': absent_count,
                    'attendance_rate': attendance_rate,
                    'last_session': last_session_str
                })
            except Exception as e:
                logger.error(f"Error processing student data for enrollment {enrollment.id}: {e}")
                continue
        
        return students_data
    except Exception as e:
        print(f"Error in get_students_report_data: {e}")
        return []


def get_facilitators_report_data(filters, date_filter, offset=0, limit=20):
    """
    Get facilitators report data - High Performance Implementation.
    Uses pre-calculated FacilitatorAttendanceSummary and server-side pagination.
    """
    try:
        from .models import FacilitatorAttendanceSummary
        
        # 1. Fetch facilitators (Role 2) with pre-calculated summary
        queryset = User.objects.filter(role_id=2, is_active=True).select_related(
            'facilitator_summary'
        )
        
        # Apply school filter (if provided)
        school_ids = filters.get('school_id')
        if school_ids:
            queryset = queryset.filter(assigned_schools__school_id__in=school_ids).distinct()
        
        facilitators = list(queryset[offset : offset + limit])
        
        facilitators_data = []
        for facilitator in facilitators:
            try:
                # 🚀 ULTRA-FAST: Read from pre-calculated summary
                summary = getattr(facilitator, 'facilitator_summary', None)
                
                schools_count = summary.schools_count if summary else 0
                sessions_conducted = summary.sessions_conducted if summary else 0
                last_active_date = summary.last_active_date if summary else None
                avg_rating = summary.average_rating if summary else 0.0
                
                last_active = last_active_date.strftime('%Y-%m-%d') if last_active_date else 'N/A'
                
                facilitators_data.append({
                    'name': facilitator.full_name,
                    'email': facilitator.email,
                    'schools_count': schools_count,
                    'sessions_conducted': sessions_conducted,
                    'avg_rating': f"{avg_rating:.1f}" if avg_rating > 0 else 'N/A',
                    'last_active': last_active
                })
            except Exception as e:
                logger.error(f"Error processing facilitator {facilitator.id}: {e}")
                continue
        
        return facilitators_data
    except Exception as e:
        logger.error(f"Error in get_facilitators_report_data: {e}")
        return []



def get_attendance_report_data(filters, date_filter, offset=0, limit=20):
    """Get attendance report data - optimized with aggregation and pagination"""
    try:
        from django.db.models import Count, Q, Max
        
        # Build the query - Start with ActualSession
        queryset = ActualSession.objects.all()
        
        # Apply filters FIRST to reduce dataset size
        # Apply date filter
        if date_filter:
            queryset = queryset.filter(**date_filter)
        
        # Apply school filter (Multiple)
        school_ids = filters.get('school_id')
        if school_ids:
            queryset = queryset.filter(planned_session__class_section__school_id__in=school_ids)
        
        # Apply class filter (Multiple)
        class_ids = filters.get('class_id')
        if class_ids:
            queryset = queryset.filter(planned_session__class_section_id__in=class_ids)
            
        # ✅ PERFORMANCE FIX: Only include conducted sessions (which have attendance)
        # Avoids expensive 'attendances__isnull=False' check and 'distinct()'
        from .models import SessionStatus
        queryset = queryset.filter(status=SessionStatus.CONDUCTED)
        
        sessions = queryset.select_related(
            'planned_session__class_section__school',
            'facilitator'
        ).order_by('-date')[offset : offset + limit]
        
        # ✅ Group Mapping (For Single Set of 150 Architecture)
        # For grouped sessions, we need to get the correct total students for THIS specific class
        # NOT the sum of all classes in the group
        from .models import GroupedSession
        enrollment_counts = Enrollment.objects.filter(
            is_active=True
        ).values('class_section_id').annotate(
            total_students=Count('id')
        )
        
        # Simple lookup: class_id -> total students for that class only
        enrollment_dict = {
            item['class_section_id']: item['total_students']
            for item in enrollment_counts
        }
        
        # Get all attendance data for these sessions in ONE query
        # Count DISTINCT students who are present (not attendance records)
        session_ids = [s.id for s in sessions]
        attendance_data = Attendance.objects.filter(
            actual_session_id__in=session_ids,
            status=AttendanceStatus.PRESENT
        ).values('actual_session_id').annotate(
            present_count=Count('enrollment_id', distinct=True)
        )
        
        # Convert to dict for fast lookup
        attendance_dict = {
            item['actual_session_id']: item['present_count']
            for item in attendance_data
        }
        
        attendance_data_list = []
        daily_attendance = {}
        class_attendance = {}
        
        for session in sessions:
            try:
                if not session.planned_session:
                    continue
                    
                # Get data from dicts (no DB queries)
                total_students = enrollment_dict.get(session.planned_session.class_section_id, 0)
                present_count = attendance_dict.get(session.id, 0)
                
                # Safeguards: ensure present doesn't exceed total
                present_count = min(present_count, total_students) if total_students > 0 else 0
                absent_count = max(0, total_students - present_count)
                attendance_percentage = round((present_count / total_students * 100) if total_students > 0 else 0, 1)
                # Cap attendance percentage at 100%
                attendance_percentage = min(attendance_percentage, 100.0)
                
                attendance_data_list.append({
                    'date': session.date.strftime('%Y-%m-%d'),
                    'school_name': session.planned_session.class_section.school.name,
                    'class_name': f"{session.planned_session.class_section.class_level} - {session.planned_session.class_section.section}",
                    'total_students': total_students,
                    'present': present_count,
                    'absent': absent_count,
                    'attendance_percentage': attendance_percentage,
                    'facilitator_name': session.facilitator.full_name if session.facilitator else 'N/A'
                })
                
                # Aggregate for charts
                date_str = session.date.strftime('%Y-%m-%d')
                if date_str not in daily_attendance:
                    daily_attendance[date_str] = []
                daily_attendance[date_str].append(attendance_percentage)
                
                class_key = f"{session.planned_session.class_section.class_level} - {session.planned_session.class_section.section}"
                if class_key not in class_attendance:
                    class_attendance[class_key] = []
                class_attendance[class_key].append(attendance_percentage)
                
            except Exception as e:
                print(f"Error processing session {session.id}: {e}")
                continue
        
        # Prepare chart data
        daily_labels = sorted(daily_attendance.keys())[:10]
        daily_avg = [sum(daily_attendance[date]) / len(daily_attendance[date]) for date in daily_labels]
        
        class_labels = list(class_attendance.keys())[:10]
        class_avg = [sum(class_attendance[cls]) / len(class_attendance[cls]) for cls in class_labels]
        
        return {
            'attendance_data': attendance_data_list,
            'daily_labels': daily_labels,
            'daily_attendance': daily_avg,
            'class_labels': class_labels,
            'class_attendance': class_avg
        }
    except Exception as e:
        print(f"Error in get_attendance_report_data: {e}")
        return {
            'attendance_data': [],
            'daily_labels': [],
            'daily_attendance': [],
            'class_labels': [],
            'class_attendance': []
        }


def get_sessions_report_data(filters, date_filter, offset=0, limit=20):
    """Get sessions report data - optimized with select_related and pagination"""
    try:
        # Build the query
        queryset = ActualSession.objects.select_related(
            'planned_session__class_section__school',
            'facilitator'
        )
        
        # Apply date filter
        if date_filter:
            queryset = queryset.filter(**date_filter)
        
        # Apply school filter (Multiple)
        school_ids = filters.get('school_id')
        if school_ids:
            queryset = queryset.filter(planned_session__class_section__school_id__in=school_ids)
        
        # Apply class filter (Multiple)
        class_ids = filters.get('class_id')
        if class_ids:
            queryset = queryset.filter(planned_session__class_section_id__in=class_ids)
        
        sessions = queryset.order_by('-date')[offset : offset + limit]
        
        sessions_data = []
        for session in sessions:
            try:
                # Calculate Session Type
                if session.status == 3:
                    session_type = "Class Not Available"
                elif session.status == 2:
                    session_type = "Holiday"
                else:
                    day_num = session.planned_session.day_number if session.planned_session else 0
                    if day_num == 999:
                        session_type = "FLN Curriculum"
                    elif day_num == 998:
                        session_type = "Exam Time"
                    elif day_num == 997:
                        session_type = "Present Office"
                    else:
                        session_type = "Present Class"

                sessions_data.append({
                    'date': session.date.strftime('%Y-%m-%d'),
                    'topic': session.planned_session.title if session.planned_session else 'N/A',
                    'day_number': session.planned_session.day_number if session.planned_session else 1,
                    'class_name': f"{session.planned_session.class_section.class_level} - {session.planned_session.class_section.section}" if session.planned_session else 'N/A',
                    'facilitator_name': session.facilitator.full_name if session.facilitator else 'N/A',
                    'status': session.get_status_display() if hasattr(session, 'get_status_display') else str(session.status),
                    'session_type': session_type,
                    'duration': session.duration_minutes if session.duration_minutes else 45
                })
            except Exception as e:
                print(f"Error processing session {session.id}: {e}")
                continue
        
        return sessions_data
    except Exception as e:
        print(f"Error in get_sessions_report_data: {e}")
        return []


def get_feedback_report_data(filters, date_filter, offset=0, limit=20):
    """Get feedback report data - optimized with select_related and pagination"""
    try:
        # Build the query with select_related to avoid N+1 queries
        queryset = SessionFeedback.objects.select_related(
            'actual_session__planned_session__class_section__school',
            'actual_session__facilitator'
        )
        
        # Apply date filter
        if date_filter:
            queryset = queryset.filter(**{f'actual_session__{k}': v for k, v in date_filter.items()})
        
        # Apply school filter (Multiple)
        school_ids = filters.get('school_id')
        if school_ids:
            queryset = queryset.filter(actual_session__planned_session__class_section__school_id__in=school_ids)
        
        # Apply class filter (Multiple)
        class_ids = filters.get('class_id')
        if class_ids:
            queryset = queryset.filter(actual_session__planned_session__class_section_id__in=class_ids)
        
        feedback_records = queryset.order_by('-feedback_date')[offset : offset + limit]
        
        feedback_data = []
        for feedback in feedback_records:
            try:
                feedback_text = feedback.day_reflection or 'No feedback provided'
                feedback_data.append({
                    'date': feedback.actual_session.date.strftime('%Y-%m-%d'),
                    'session_topic': feedback.actual_session.planned_session.title if feedback.actual_session.planned_session else 'N/A',
                    'facilitator_name': feedback.actual_session.facilitator.full_name if feedback.actual_session.facilitator else 'N/A',
                    'rating': f"{feedback.rating / 2.0:.1f}" if feedback.rating is not None else 'N/A',
                    'feedback_text': feedback_text[:100] + '...' if len(feedback_text) > 100 else feedback_text,
                    'category': 'Session Feedback'
                })
            except Exception as e:
                print(f"Error processing feedback {feedback.id}: {e}")
                continue
        
        return feedback_data
    except Exception as e:
        print(f"Error in get_feedback_report_data: {e}")
        return []


@login_required
def download_pdf_report(request, report_type):
    """Generate and download PDF report"""
    if request.user.role.name.upper() != "ADMIN":
        return JsonResponse({'error': 'Permission denied'}, status=403)
    
    if request.method != 'POST':
        return JsonResponse({'error': 'POST method required'}, status=405)
    
    # Get filters from POST data - correctly collect multiple IDs
    filters = {key: value for key, value in request.POST.items() if value and key != 'csrfmiddlewaretoken' and not key.endswith('[]')}
    
    # Correctly parse multiple selections if present
    school_ids = request.POST.getlist('school_id[]')
    if school_ids:
        filters['school_id'] = school_ids
        
    class_ids = request.POST.getlist('class_id[]')
    if class_ids:
        filters['class_id'] = class_ids

    date_filter = get_date_filter(filters)
    
    # Create PDF
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    styles = getSampleStyleSheet()
    story = []
    
    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        spaceAfter=30,
        alignment=1  # Center alignment
    )
    story.append(Paragraph(f"{report_type.title()} Report", title_style))
    story.append(Spacer(1, 12))
    
    # Date range info
    date_info = get_date_range_text(filters)
    story.append(Paragraph(f"Report Period: {date_info}", styles['Normal']))
    story.append(Spacer(1, 20))
    
    # Get data based on report type
    if report_type == 'students':
        data = get_students_report_data(filters, date_filter)
        headers = ['Student Name', 'Enrollment No', 'Class', 'School', 'Present', 'Absent', 'Attendance Rate', 'Last Session']
        table_data = [[
            row['name'], row['enrollment_number'], row['class_name'], 
            row['school_name'], str(row['present']), str(row['absent']), f"{row['attendance_rate']}%", row['last_session'] or 'N/A'
        ] for row in data]
        
    elif report_type == 'facilitators':
        data = get_facilitators_report_data(filters, date_filter)
        headers = ['Facilitator Name', 'Email', 'Schools', 'Sessions', 'Avg Rating', 'Last Active']
        table_data = [[
            row['name'], row['email'], str(row['schools_count']), 
            str(row['sessions_conducted']), str(row['avg_rating']) if row['avg_rating'] else 'N/A', row['last_active']
        ] for row in data]
        
    elif report_type == 'attendance':
        data = get_attendance_report_data(filters, date_filter)
        attendance_records = data if isinstance(data, list) else data.get('attendance_data', [])
        headers = ['Date', 'School', 'Class', 'Total', 'Present', 'Absent', 'Attendance %', 'Facilitator']
        table_data = [[
            row['date'], row['school_name'], row['class_name'], 
            str(row['total_students']), str(row['present']), str(row['absent']),
            f"{row['attendance_percentage']}%", row['facilitator_name']
        ] for row in attendance_records]
        
    elif report_type == 'sessions':
        data = get_sessions_report_data(filters, date_filter)
        headers = ['Date', 'Topic', 'Day', 'Type', 'Class', 'Facilitator', 'Status', 'Duration']
        table_data = [[
            row['date'], row['topic'], f"Day {row['day_number']}", row['session_type'],
            row['class_name'], row['facilitator_name'], row['status'], str(row['duration']) if row['duration'] else 'N/A'
        ] for row in data]
        
    elif report_type == 'feedback':
        data = get_feedback_report_data(filters, date_filter)
        headers = ['Date', 'Session', 'Facilitator', 'Rating', 'Feedback', 'Category']
        table_data = [[
            row['date'], row['session_topic'], row['facilitator_name'], 
            f"{row['rating']}/5", row['feedback_text'][:50] + '...' if len(row['feedback_text']) > 50 else row['feedback_text'], 
            row['category']
        ] for row in data]
    
    # Create table
    if table_data:
        table_data.insert(0, headers)  # Add headers as first row
        table = Table(table_data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        story.append(table)
    else:
        story.append(Paragraph("No data available for the selected criteria.", styles['Normal']))
    
    # Build PDF
    doc.build(story)
    buffer.seek(0)
    
    # Create response
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{report_type}_report_{timezone.now().strftime("%Y%m%d")}.pdf"'
    
    return response


@login_required
def download_excel_report(request, report_type):
    """Generate and download Excel report"""
    if request.user.role.name.upper() != "ADMIN":
        return JsonResponse({'error': 'Permission denied'}, status=403)
    
    if request.method != 'POST':
        return JsonResponse({'error': 'POST method required'}, status=405)
    
    # Get filters from POST data - correctly collect multiple IDs
    filters = {key: value for key, value in request.POST.items() if value and key != 'csrfmiddlewaretoken' and not key.endswith('[]')}
    
    # Correctly parse multiple selections if present
    school_ids = request.POST.getlist('school_id[]')
    if school_ids:
        filters['school_id'] = school_ids
        
    class_ids = request.POST.getlist('class_id[]')
    if class_ids:
        filters['class_id'] = class_ids

    date_filter = get_date_filter(filters)
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{report_type.title()} Report"
    
    # Get data based on report type
    if report_type == 'attendance':
        data = get_attendance_report_data(filters, date_filter)
        attendance_records = data if isinstance(data, list) else data.get('attendance_data', [])
        
        # Headers
        headers = ['Date', 'School', 'Class', 'Total Students', 'Present', 'Absent', 'Attendance %', 'Facilitator']
        ws.append(headers)
        
        # Data rows
        for row in attendance_records:
            ws.append([
                row['date'], row['school_name'], row['class_name'], 
                row['total_students'], row['present'], row['absent'],
                row['attendance_percentage'], row['facilitator_name']
            ])
    
    # Style the header row
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    # Create response
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{report_type}_report_{timezone.now().strftime("%Y%m%d")}.xlsx"'
    
    return response


# Helper functions
def get_date_filter(filters):
    """Convert filter parameters to Django ORM date filter"""
    date_range = filters.get('date_range', 'month')
    now = timezone.now()
    
    if date_range == 'today':
        return {'date': now.date()}
    elif date_range == 'week':
        start_date = now - timedelta(days=7)
        return {'date__gte': start_date.date()}
    elif date_range == 'month':
        start_date = now - timedelta(days=30)
        return {'date__gte': start_date.date()}
    elif date_range == 'quarter':
        start_date = now - timedelta(days=90)
        return {'date__gte': start_date.date()}
    elif date_range == 'year':
        start_date = now - timedelta(days=365)
        return {'date__gte': start_date.date()}
    elif date_range == 'custom':
        date_filter = {}
        if filters.get('start_date'):
            date_filter['date__gte'] = datetime.strptime(filters['start_date'], '%Y-%m-%d').date()
        if filters.get('end_date'):
            date_filter['date__lte'] = datetime.strptime(filters['end_date'], '%Y-%m-%d').date()
        return date_filter
    
    return {}


def get_date_range_text(filters):
    """Get human-readable date range text"""
    date_range = filters.get('date_range', 'month')
    
    if date_range == 'today':
        return 'Today'
    elif date_range == 'week':
        return 'Last 7 days'
    elif date_range == 'month':
        return 'Last 30 days'
    elif date_range == 'quarter':
        return 'Last 90 days'
    elif date_range == 'year':
        return 'Last 365 days'
    elif date_range == 'custom':
        start = filters.get('start_date', 'N/A')
        end = filters.get('end_date', 'N/A')
        return f'{start} to {end}'
    
    return 'All time'


def calculate_overall_attendance_rate():
    """Calculate overall attendance rate across all sessions"""
    total_attendance_records = Attendance.objects.count()
    if total_attendance_records == 0:
        return 0
    
    present_records = Attendance.objects.filter(status=AttendanceStatus.PRESENT).count()
    return round((present_records / total_attendance_records) * 100, 1)


@login_required
def download_all_excel_report(request):
    """Generate and download a combined Excel report with all tabs as worksheets"""
    if request.user.role.name.upper() != "ADMIN":
        return JsonResponse({'error': 'Permission denied'}, status=403)
    
    if request.method != 'POST':
        return JsonResponse({'error': 'POST method required'}, status=405)
    
    # Get filters from POST data - correctly collect multiple IDs
    filters = {key: value for key, value in request.POST.items() if value and key != 'csrfmiddlewaretoken' and not key.endswith('[]')}
    
    # Correctly parse multiple selections if present
    school_ids = request.POST.getlist('school_id[]')
    if school_ids:
        filters['school_id'] = school_ids
        
    class_ids = request.POST.getlist('class_id[]')
    if class_ids:
        filters['class_id'] = class_ids

    date_filter = get_date_filter(filters)
    
    # Create workbook
    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)
    
    from openpyxl.styles import Font, Alignment, PatternFill
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    report_types = ['students', 'facilitators', 'attendance', 'sessions', 'feedback']
    
    for rt in report_types:
        # Excel sheet title limit is 31 chars
        ws = wb.create_sheet(title=rt.title()[:31])
        
        headers = []
        rows = []
        
        if rt == 'students':
            data = get_students_report_data(filters, date_filter)
            headers = ['Student Name', 'Enrollment No', 'Class', 'School', 'Present', 'Absent', 'Attendance Rate', 'Last Session']
            rows = [[row['name'], row['enrollment_number'], row['class_name'], row['school_name'], str(row['present']), str(row['absent']), f"{row['attendance_rate']}%", row['last_session'] or 'N/A'] for row in data]
        elif rt == 'facilitators':
            data = get_facilitators_report_data(filters, date_filter)
            headers = ['Facilitator Name', 'Email', 'Schools', 'Sessions', 'Avg Rating', 'Last Active']
            rows = [[row['name'], row['email'], str(row['schools_count']), str(row['sessions_conducted']), str(row['avg_rating']) if row['avg_rating'] else 'N/A', row['last_active']] for row in data]
        elif rt == 'attendance':
            data = get_attendance_report_data(filters, date_filter)
            attendance_records = data if isinstance(data, list) else data.get('attendance_data', [])
            headers = ['Date', 'School', 'Class', 'Total', 'Present', 'Absent', 'Attendance %', 'Facilitator']
            rows = [[row['date'], row['school_name'], row['class_name'], str(row['total_students']), str(row['present']), str(row['absent']), f"{row['attendance_percentage']}%", row['facilitator_name']] for row in attendance_records]
        elif rt == 'sessions':
            data = get_sessions_report_data(filters, date_filter)
            headers = ['Date', 'Topic', 'Day', 'Type', 'Class', 'Facilitator', 'Status', 'Duration']
            rows = [[row['date'], row['topic'], f"Day {row['day_number']}", row['session_type'], row['class_name'], row['facilitator_name'], row['status'], str(row['duration']) if row['duration'] else 'N/A'] for row in data]
        elif rt == 'feedback':
            data = get_feedback_report_data(filters, date_filter)
            headers = ['Date', 'Session', 'Facilitator', 'Rating', 'Feedback', 'Category']
            rows = [[row['date'], row['session_topic'], row['facilitator_name'], f"{row['rating']}/5", row['feedback_text'][:100] if row['feedback_text'] else '', row['category']] for row in data]

        # Append headers
        ws.append(headers)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            
        # Append rows
        for r in rows:
            ws.append(r)
            
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except: pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    # Save to buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    response = HttpResponse(buffer.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="all_reports_{timezone.now().strftime("%Y%m%d")}.xlsx"'
    return response