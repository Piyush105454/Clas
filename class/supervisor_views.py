# Supervisor Views - Complete Management Interface
import uuid
import logging
from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse
from django.contrib.auth import get_user_model
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.db.models import Count, Q, Prefetch, Exists, OuterRef, Avg, Sum, Max, Min
from django.core.cache import cache
from django.db import transaction
from django.utils import timezone
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from io import BytesIO
from .models import (
    Role, School, ClassSection, FacilitatorSchool, PlannedSession, 
    DateType, Cluster, Enrollment, GroupedSession, User as CoreUser,
    ActualSession, Attendance, SessionStatus, AttendanceStatus, CalendarDate
)
from .forms import AddUserForm, EditUserForm, AddSchoolForm, EditSchoolForm, ClassSectionForm, AssignFacilitatorForm, ClusterForm
from .decorators import supervisor_required, admin_required

User = get_user_model()
logger = logging.getLogger(__name__)

# =====================================================
# Helper Functions for Grouped Sessions
# =====================================================
def initialize_grouped_session_plans(classes, grouped_session_id):
    """
    Initialize 150-day session plans for grouped classes.
    
    - Ensures exactly 150 master sessions exist for the primary class (classes[0])
    - Deletes physical sessions for secondary classes (they share the primary's sessions)
    - Links the 150 master sessions to the grouped_session_id
    """
    if not classes:
        return {'success': False, 'error': 'No classes provided'}
    
    try:
        from .signals_optimization import silence_signals
        with transaction.atomic(), silence_signals():
            classes = list(classes)
            primary_class = classes[0]
            total_sessions_created = 0
            
            # Step 1: Handle Primary Class (Master Record Holder)
            all_ps = list(PlannedSession.objects.filter(class_section=primary_class).order_by('day_number', 'created_at'))
            seen_days = set()
            to_delete = []
            valid_sessions = []
            
            for ps in all_ps:
                if ps.day_number in seen_days or ps.day_number > 150:
                    to_delete.append(ps.id)
                else:
                    seen_days.add(ps.day_number)
                    valid_sessions.append(ps)
            
            if to_delete:
                PlannedSession.objects.filter(id__in=to_delete).delete()
            
            # If sessions are missing for primary class, create them
            missing_days = set(range(1, 151)) - seen_days
            if missing_days:
                sessions_to_create = []
                for day_num in sorted(list(missing_days)):
                    sessions_to_create.append(PlannedSession(
                        class_section=primary_class,
                        day_number=day_num,
                        title=f"Day {day_num} Session",
                        description=f"Initial session setup",
                        sequence_position=day_num,
                        is_required=True,
                        is_active=True,
                        grouped_session_id=grouped_session_id
                    ))
                PlannedSession.objects.bulk_create(sessions_to_create)
                total_sessions_created += len(sessions_to_create)
            
            # Ensure all primary sessions have the grouped ID
            PlannedSession.objects.filter(class_section=primary_class, day_number__lte=150).update(grouped_session_id=grouped_session_id)
            
            # Step 2: Clear Secondary Classes (They leverage the Primary's sessions)
            for secondary_class in classes[1:]:
                PlannedSession.objects.filter(class_section=secondary_class).delete()

            # Create a GroupedSession record to permanently track this grouping
            grouped_session, created = GroupedSession.objects.get_or_create(
                grouped_session_id=grouped_session_id,
                defaults={
                    'name': f"Grouped: {', '.join([c.display_name for c in classes])}",
                    'description': f"Permanent grouping of {len(classes)} classes sharing 150 sessions"
                }
            )
            
            # Add all classes to the GroupedSession
            grouped_session.class_sections.set(classes)
            
            return {
                'success': True,
                'message': f'Initialized grouped session plan for {len(classes)} classes',
                'classes_count': len(classes),
                'sessions_created': total_sessions_created,
                'grouped_session_id': str(grouped_session_id)
            }
    
    except Exception as e:
        logger.error(f"Error initializing grouped session: {str(e)}")
        return {'success': False, 'error': str(e)}

# =====================================================
# Permission Decorator for Supervisor
# =====================================================
def supervisor_required(view_func):
    def wrapper(request, *args, **kwargs):
        if request.user.role.name.upper() != "SUPERVISOR":
            messages.error(request, "You do not have permission to access this page.")
            return redirect("no_permission")
        return view_func(request, *args, **kwargs)
    return wrapper

# =====================================================
# Dashboard
# =====================================================
@login_required
@supervisor_required
def supervisor_dashboard(request):
    """Supervisor Dashboard - Overview of all managed resources"""
    
    from django.db.models import Count, Q
    
    # Use aggregation instead of count() for better performance
    stats = User.objects.aggregate(
        active_facilitators=Count('id', filter=Q(role__name__iexact="FACILITATOR", is_active=True))
    )
    
    # Batch queries for counts
    school_stats = School.objects.aggregate(
        total_schools=Count('id'),
        active_schools=Count('id', filter=Q(status=1))
    )
    
    class_stats = ClassSection.objects.aggregate(
        total_classes=Count('id'),
        active_classes=Count('id', filter=Q(is_active=True))
    )
    
    # Batch query: Get recent users and schools with select_related
    recent_users = list(User.objects.all().select_related('role').order_by("-created_at")[:5])
    recent_schools = list(School.objects.all().order_by("-created_at")[:5])
    
    logger.info(f"Dashboard - Active Facilitators: {stats['active_facilitators']}, Total Schools: {school_stats['total_schools']}, Total Classes: {class_stats['total_classes']}")
    
    context = {
        'total_schools': school_stats['total_schools'],
        'active_schools': school_stats['active_schools'],
        'total_classes': class_stats['total_classes'],
        'active_classes': class_stats['active_classes'],
        'active_facilitators': stats['active_facilitators'],
        'recent_users': recent_users,
        'recent_schools': recent_schools,
    }
    
    return render(request, "supervisor/dashboard.html", context)

# =====================================================
# User Management
# =====================================================
@login_required
@supervisor_required
def supervisor_users_list(request):
    """List all users with filtering"""
    
    users = User.objects.all().select_related('role').order_by("-created_at")
    
    # Filter by role
    role_filter = request.GET.get('role')
    if role_filter:
        users = users.filter(role__id=role_filter)
    
    # Filter by status
    status_filter = request.GET.get('status')
    if status_filter == 'active':
        users = users.filter(is_active=True)
    elif status_filter == 'inactive':
        users = users.filter(is_active=False)
    
    # Batch query: Get all roles at once
    roles = Role.objects.all()
    
    context = {
        'users': users,
        'roles': roles,
        'selected_role': role_filter,
        'selected_status': status_filter,
    }
    
    return render(request, "supervisor/users/list.html", context)

@login_required
@supervisor_required
def supervisor_user_create(request):
    """Create new user with role assignment"""
    
    if request.method == "POST":
        form = AddUserForm(request.POST)
        if form.is_valid():
            user = form.save(commit=False)
            user.set_password(form.cleaned_data["password"])
            user.save()
            messages.success(request, f"User '{user.full_name}' created successfully!")
            return redirect("supervisor_users_list")
    else:
        form = AddUserForm()
    
    return render(request, "supervisor/users/create.html", {"form": form})

@login_required
@supervisor_required
def supervisor_user_edit(request, user_id):
    """Edit user details"""
    
    user = get_object_or_404(User, id=user_id)
    
    if request.method == "POST":
        form = EditUserForm(request.POST, instance=user)
        if form.is_valid():
            form.save()
            messages.success(request, f"User '{user.full_name}' updated successfully!")
            return redirect("supervisor_users_list")
    else:
        form = EditUserForm(instance=user)
    
    return render(request, "supervisor/users/edit.html", {
        "form": form,
        "user": user
    })

@login_required
@supervisor_required
def supervisor_user_delete(request, user_id):
    """Delete user"""
    
    user = get_object_or_404(User, id=user_id)
    
    if request.method == "POST":
        user_name = user.full_name
        user.delete()
        messages.success(request, f"User '{user_name}' deleted successfully!")
        return redirect("supervisor_users_list")
    
    return render(request, "supervisor/users/delete_confirm.html", {"user": user})

# =====================================================
# School Management
# =====================================================
@login_required
@supervisor_required
def supervisor_schools_list(request):
    """List all schools - OPTIMIZED with database-level filtering"""
    
    # Get status filter
    status_filter = request.GET.get('status')
    
    # Build query with prefetch
    query = School.objects.prefetch_related(
        Prefetch(
            'class_sections',
            queryset=ClassSection.objects.filter(is_active=True)
        ),
        Prefetch(
            'facilitators',
            queryset=FacilitatorSchool.objects.select_related('facilitator').filter(is_active=True)
        )
    ).annotate(
        total_classes=Count('class_sections', filter=Q(class_sections__is_active=True), distinct=True),
        total_students=Count('class_sections__enrollments', 
                           filter=Q(class_sections__enrollments__is_active=True),
                           distinct=True),
        active_facilitators=Count('facilitators', 
                                filter=Q(facilitators__is_active=True),
                                distinct=True)
    )
    
    # Apply status filter at database level (not in Python)
    if status_filter:
        query = query.filter(status=int(status_filter))
    
    # Order and execute query
    schools = query.order_by("-created_at")
    
    context = {
        'schools': schools,
        'selected_status': status_filter,
    }
    
    return render(request, "supervisor/schools/list.html", context)

@login_required
@supervisor_required
def supervisor_school_create(request):
    """Create new school"""
    
    if request.method == "POST":
        form = AddSchoolForm(request.POST, request.FILES)
        if form.is_valid():
            school = form.save()
            cache_key = f"supervisor_schools_list_{request.user.id}"
            cache.delete(cache_key)
            messages.success(request, f"School '{school.name}' created successfully!")
            return redirect("supervisor_schools_list")
    else:
        form = AddSchoolForm()
    
    return render(request, "supervisor/schools/create.html", {"form": form})

@login_required
@supervisor_required
def supervisor_school_edit(request, school_id):
    """Edit school details"""
    
    school = get_object_or_404(School, id=school_id)
    
    if request.method == "POST":
        form = EditSchoolForm(request.POST, request.FILES, instance=school)
        if form.is_valid():
            school = form.save()
            cache_key = f"supervisor_schools_list_{request.user.id}"
            cache.delete(cache_key)
            messages.success(request, f"School '{school.name}' updated successfully!")
            return redirect("supervisor_schools_list")
    else:
        form = EditSchoolForm(instance=school)
    
    # Get existing blocks for the district
    existing_blocks = []
    if school.district:
        existing_blocks = list(School.objects.filter(
            district__iexact=school.district
        ).values_list('block', flat=True).distinct().order_by('block'))
    
    return render(request, "supervisor/schools/edit.html", {
        "form": form,
        "school": school,
        "existing_blocks": existing_blocks
    })

@login_required
@supervisor_required
def supervisor_school_detail(request, school_id):
    """View school details"""
    
    school = get_object_or_404(School, id=school_id)
    classes = ClassSection.objects.filter(school=school).order_by("class_level", "section")
    facilitators = FacilitatorSchool.objects.filter(school=school).select_related("facilitator")
    
    context = {
        'school': school,
        'classes': classes,
        'facilitators': facilitators,
    }
    
    return render(request, "supervisor/schools/detail.html", context)

@login_required
@supervisor_required
def supervisor_school_delete(request, school_id):
    """Delete school"""
    
    school = get_object_or_404(School, id=school_id)
    
    if request.method == "POST":
        school_name = school.name
        school.delete()
        cache_key = f"supervisor_schools_list_{request.user.id}"
        cache.delete(cache_key)
        messages.success(request, f"School '{school_name}' deleted successfully!")
        return redirect("supervisor_schools_list")
    
    return render(request, "supervisor/schools/delete_confirm.html", {"school": school})

# =====================================================
# Class Management
# =====================================================
@login_required
@supervisor_required
def supervisor_classes_list(request):
    """List classes - OPTIMIZED: Only load classes when school is selected"""
    from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
    
    # Get school filter
    school_filter = request.GET.get('school')
    page = request.GET.get('page', 1)
    
    # Only load classes if school is selected
    classes_page = None
    paginator = None
    
    if school_filter:
        # Build query with school filter at database level
        query = ClassSection.objects.select_related('school').prefetch_related(
            Prefetch(
                'enrollments',
                queryset=Enrollment.objects.filter(is_active=True)
            ),
            Prefetch(
                'planned_sessions',
                queryset=PlannedSession.objects.filter(is_active=True)
            )
        ).filter(school__id=school_filter).order_by("class_level", "section")
        
        # Paginate: 50 classes per page
        paginator = Paginator(query, 50)
        try:
            classes_page = paginator.page(page)
        except PageNotAnInteger:
            classes_page = paginator.page(1)
        except EmptyPage:
            classes_page = paginator.page(paginator.num_pages)
    
    # Batch query: Get all schools at once
    schools = cache.get("supervisor_schools_all")
    if schools is None:
        schools = School.objects.all().order_by('name')
        cache.set("supervisor_schools_all", schools, 300)
    
    context = {
        'classes': classes_page,
        'paginator': paginator,
        'schools': schools,
        'selected_school': school_filter,
    }
    
    return render(request, "supervisor/classes/list.html", context)

# =====================================================
# Reports & Analytics
# =====================================================
def _reports_dashboard_logic(request, base_template='supervisor/shared/base.html'):
    """Core logic for reports and analytics dashboard (undecorated)"""
    """Reports and analytics dashboard"""
    
    from django.db.models import Count, Q
    
    # Use aggregation for all statistics
    stats = User.objects.aggregate(
        total_users=Count('id')
    )
    
    total_schools = School.objects.count()
    total_classes = ClassSection.objects.count()
    
    # Batch query: Get user breakdown by role
    user_by_role = User.objects.values('role__name').annotate(count=Count('id'))
    
    # Batch query: Get schools by status
    schools_by_status = School.objects.values('status').annotate(count=Count('id'))
    
    context = {
        'total_users': stats['total_users'],
        'total_schools': total_schools,
        'total_classes': total_classes,
        'user_by_role': user_by_role,
        'schools_by_status': schools_by_status,
        'base_template': base_template,
    }
    
    return render(request, "supervisor/reports/dashboard.html", context)


@login_required
@supervisor_required
def supervisor_reports_dashboard(request, base_template='supervisor/shared/base.html'):
    """Supervisor entry point for reports dashboard"""
    return _reports_dashboard_logic(request, base_template=base_template)


@login_required
@admin_required
def admin_reports_dashboard(request):
    """Admin entry point for reports dashboard (uses admin layout)"""
    return _reports_dashboard_logic(request, base_template='admin/shared/base.html')


def _feedback_analytics_logic(request, base_template='supervisor/shared/base.html'):
    """Core logic for feedback analytics (undecorated)"""
    context = {
        'base_template': base_template,
    }
    return render(request, "supervisor/reports/feedback.html", context)


@login_required
@supervisor_required
def supervisor_feedback_analytics(request, base_template='supervisor/shared/base.html'):
    """Supervisor entry point for feedback analytics"""
    return _feedback_analytics_logic(request, base_template=base_template)


@login_required
@admin_required
def admin_feedback_dashboard(request):
    """Admin entry point for feedback analytics (uses admin layout)"""
    return _feedback_analytics_logic(request, base_template='admin/shared/base.html')

# =====================================================
# Settings
# =====================================================
@login_required
@supervisor_required
def supervisor_settings(request):
    """Supervisor settings"""
    
    return render(request, "supervisor/settings.html", {})

# =====================================================
# AJAX Endpoints
# =====================================================
@csrf_exempt
@login_required
@supervisor_required
def get_blocks_by_district(request):
    """Get blocks for a district"""
    district = request.GET.get('district', '')
    
    if not district:
        return JsonResponse({'blocks': []})
    
    # Get unique blocks from schools in this district
    blocks = School.objects.filter(
        district__iexact=district
    ).values_list('block', flat=True).distinct().order_by('block')
    
    return JsonResponse({'blocks': list(blocks)})

@csrf_exempt
@login_required
@supervisor_required
def get_schools_by_block(request):
    """Get schools for a block"""
    district = request.GET.get('district', '')
    block = request.GET.get('block', '')
    
    if not district or not block:
        return JsonResponse({'schools': []})
    
    # Get schools in this district and block
    schools = School.objects.filter(
        district__iexact=district,
        block__iexact=block
    ).values('id', 'name', 'latitude', 'longitude').order_by('name')
    
    return JsonResponse({'schools': list(schools)})

@csrf_exempt
@login_required
@supervisor_required
def get_all_schools(request):
    """Get all schools with coordinates for map display"""
    schools = School.objects.exclude(
        latitude=28.7041,
        longitude=77.1025
    ).values('id', 'name', 'district', 'block', 'latitude', 'longitude').order_by('name')
    
    return JsonResponse({'schools': list(schools)})

@csrf_exempt
@login_required
@supervisor_required
def supervisor_create_user_ajax(request):
    """AJAX endpoint for creating users"""
    
    if request.method == "POST":
        full_name = request.POST.get("full_name")
        email = request.POST.get("email")
        password = request.POST.get("password")
        role_id = request.POST.get("role")
        
        if not all([full_name, email, password, role_id]):
            return JsonResponse({"success": False, "error": "All fields are required."})
        
        if User.objects.filter(email=email).exists():
            return JsonResponse({"success": False, "error": "User already exists."})
        
        try:
            role = Role.objects.get(id=role_id)
            user = User.objects.create_user(
                email=email,
                password=password,
                full_name=full_name,
                role=role
            )
            
            return JsonResponse({
                "success": True,
                "user": {
                    "id": str(user.id),
                    "full_name": user.full_name,
                    "email": user.email,
                    "role_name": role.name
                }
            })
        
        except Role.DoesNotExist:
            return JsonResponse({"success": False, "error": "Invalid role selected."})
# =====================================================
# Facilitator Management
# =====================================================
@login_required
@supervisor_required
def supervisor_facilitators_list(request):
    """Supervisor entry point for facilitators list"""
    return _facilitators_list_logic(request, base_template='supervisor/shared/base.html')


@login_required
@admin_required
def admin_view_facilitators(request):
    """Admin entry point for facilitators list (uses admin layout)"""
    return _facilitators_list_logic(request, base_template='admin/shared/base.html')


def _facilitators_list_logic(request, base_template='supervisor/shared/base.html'):
    """Core logic for facilitators list (undecorated)"""
    
    from .query_optimizations import OptimizedQueries
    from django.db.models import Count, Q
    
    # Get facilitators with prefetched schools (1 query)
    facilitators = User.objects.filter(
        role__name__iexact="FACILITATOR"
    ).select_related('role').prefetch_related(
        Prefetch(
            'assigned_schools',
            queryset=FacilitatorSchool.objects.select_related('school').filter(is_active=True)
        )
    ).annotate(
        school_count=Count('assigned_schools', filter=Q(assigned_schools__is_active=True), distinct=True)
    ).order_by("-created_at")
    
    facilitator_ids = [f.id for f in facilitators]
    
    # Get all stats in ONE query using optimized utility
    stats = OptimizedQueries.get_facilitator_stats(facilitator_ids)
    
    # Enrich facilitators with stats
    facilitator_list = []
    for facilitator in facilitators:
        facilitator.total_sessions = stats['sessions'].get(facilitator.id, 0)
        facilitator.total_strength = stats['students'].get(facilitator.id, 0)
        feedback_data = stats['feedback'].get(facilitator.id, {})
        facilitator.feedback_count = feedback_data.get('count', 0)
        facilitator.avg_rating = feedback_data.get('avg_rating', 0)
        facilitator_list.append(facilitator)
    
    # Filter by status
    status_filter = request.GET.get('status')
    if status_filter == 'active':
        facilitator_list = [f for f in facilitator_list if f.is_active]
    elif status_filter == 'inactive':
        facilitator_list = [f for f in facilitator_list if not f.is_active]
    
    context = {
        'facilitators': facilitator_list,
        'selected_status': status_filter,
    }
    
    return render(request, "supervisor/facilitators/list.html", context)

def _session_detail_logic(request, session_id, base_template='supervisor/shared/base.html'):
    """Core logic for supervisor session detail (undecorated)"""
    from .models import ActualSession, Attendance, StudentFeedback, LessonPlanUpload
    
    session = get_object_or_404(PlannedSession, id=session_id)
    
    # Get actual session if it exists
    actual_session = ActualSession.objects.filter(
        planned_session=session
    ).first()
    
    # Get attendance records (Ensure always a QuerySet)
    attendance_records = Attendance.objects.none()
    if actual_session:
        attendance_records = Attendance.objects.filter(
            actual_session=actual_session
        ).select_related('enrollment__student')
    
    # Get student feedback (Ensure always a QuerySet)
    feedback = StudentFeedback.objects.none()
    if actual_session:
        feedback = StudentFeedback.objects.filter(
            actual_session=actual_session
        )
    
    # Get lesson plan uploads for this session (from all facilitators)
    lesson_plan_uploads = LessonPlanUpload.objects.filter(
        planned_session=session
    ).order_by('-upload_date').select_related('facilitator')
    
    # Calculate statistics
    from .models import AttendanceStatus
    total_students = session.class_section.enrollments.filter(is_active=True).count()
    present_count = attendance_records.filter(status=AttendanceStatus.PRESENT).count()
    absent_count = attendance_records.filter(status=AttendanceStatus.ABSENT).count()
    leave_count = attendance_records.filter(status=AttendanceStatus.LEAVE).count()
    
    attendance_percentage = 0
    if total_students > 0:
        attendance_percentage = round((present_count / total_students) * 100, 2)
    
    # Average feedback rating
    avg_rating = 0
    if feedback.exists():
        avg_rating = feedback.aggregate(Avg('session_rating'))['session_rating__avg'] or 0
    
    # Check if current user is an admin for dynamic UI routing
    from .decorators import _is_admin
    is_admin = _is_admin(request.user)
    
    return render(request, 'supervisor/sessions/detail.html', {
        'session': session,
        'actual_session': actual_session,
        'attendance_records': attendance_records,
        'feedback': feedback,
        'lesson_plan_uploads': lesson_plan_uploads,
        'total_students': total_students,
        'present_count': present_count,
        'absent_count': absent_count,
        'leave_count': leave_count,
        'attendance_percentage': attendance_percentage,
        'avg_rating': avg_rating,
        'base_template': base_template,
        'is_admin': is_admin,
    })

@login_required
@supervisor_required
def supervisor_session_detail(request, session_id, base_template='supervisor/shared/base.html'):
    """Supervisor entry point for session detail"""
    return _session_detail_logic(request, session_id, base_template=base_template)


@login_required
@admin_required
def admin_session_detail(request, session_id):
    """Admin entry point for session detail (uses admin layout)"""
    return _session_detail_logic(request, session_id, base_template='admin/shared/base.html')


def _class_sessions_logic(request, class_id, base_template='supervisor/shared/base.html'):
    """Core logic for supervisor class sessions (undecorated)"""
    from .models import ActualSession
    
    class_section = get_object_or_404(ClassSection, id=class_id)
    
    # Get all sessions for this class
    sessions = PlannedSession.objects.filter(
        class_section=class_section
    ).order_by('day_number')
    
    # Annotate with actual session status
    sessions = sessions.prefetch_related(
        Prefetch(
            'actual_sessions',
            queryset=ActualSession.objects.select_related('planned_session')
        )
    )
    
    # Get attendance summary for each session
    session_data = []
    for session in sessions:
        actual_session = session.actual_sessions.first() if session.actual_sessions.exists() else None
        
        attendance_count = 0
        if actual_session:
            from .models import Attendance, AttendanceStatus
            attendance_count = Attendance.objects.filter(
                actual_session=actual_session,
                status=AttendanceStatus.PRESENT
            ).count()
        
        session_data.append({
            'session': session,
            'actual_session': actual_session,
            'attendance_count': attendance_count
        })
    
    return render(request, 'supervisor/sessions/class_sessions.html', {
        'class_section': class_section,
        'session_data': session_data,
        'base_template': base_template,
    })


@login_required
@supervisor_required
def supervisor_class_sessions(request, class_id, base_template='supervisor/shared/base.html'):
    """Supervisor entry point for class sessions"""
    return _class_sessions_logic(request, class_id, base_template=base_template)


@login_required
@admin_required
def admin_class_sessions(request, class_id):
    """Admin entry point for class sessions (uses admin layout)"""
    return _class_sessions_logic(request, class_id, base_template='admin/shared/base.html')

def _facilitator_detail_logic(request, facilitator_id, base_template='supervisor/shared/base.html'):
    """Core logic for facilitator profile detail (undecorated)"""
    """View facilitator profile and their work"""
    # DIRECT BYPASS: If Admin, skip decorator/middleware logic if they somehow got here
    # (This is a safety net in case the decorator logic is cached or failing)
    role_id = request.user.role.id if request.user.role else None
    if role_id == 0 or request.user.is_superuser or request.user.is_staff:
         pass # Normal flow
    
    from datetime import timedelta
    from django.utils import timezone
    
    facilitator = get_object_or_404(User, id=facilitator_id, role__name__iexact="FACILITATOR")
    
    # Get facilitator's schools
    facilitator_schools = FacilitatorSchool.objects.filter(
        facilitator=facilitator,
        is_active=True
    ).select_related('school')
    
    # Get facilitator's classes from assigned schools
    school_ids = facilitator_schools.values_list('school_id', flat=True)
    facilitator_classes = ClassSection.objects.filter(
        school_id__in=school_ids
    ).select_related('school').order_by('school__name', 'class_level', 'section')
    
    # Get date filter from request (default: last 30 days)
    date_filter = request.GET.get('date_filter', '30')
    try:
        days = int(date_filter)
    except (ValueError, TypeError):
        days = 30
    
    # Calculate date range
    end_date = timezone.localdate()
    start_date = end_date - timedelta(days=days)
    
    # Get class filter from request
    selected_class_id = request.GET.get('class_filter', None)
    
    # Get task date filter from request
    selected_task_date = request.GET.get('task_date', None)
    
    # Get facilitator's RECENT students (limit to 15)
    from django.db.models import Prefetch
    from .models import Student, Enrollment
    if selected_class_id:
        # Filter students by selected class
        facilitator_students = Student.objects.filter(
            enrollments__class_section_id=selected_class_id,
            enrollments__is_active=True
        ).distinct().order_by('-enrollments__start_date')[:15]
    else:
        # Show all students from all classes
        facilitator_students = Student.objects.filter(
            enrollments__class_section__school_id__in=school_ids,
            enrollments__is_active=True
        ).distinct().order_by('-enrollments__start_date')[:15]
    
    # Get facilitator's tasks (preparation media) - FILTERED BY SPECIFIC DATE if provided
    from .models import FacilitatorTask, SessionFeedback
    if selected_task_date:
        # Filter by specific date
        from datetime import datetime
        task_date_obj = datetime.strptime(selected_task_date, '%Y-%m-%d').date()
        facilitator_tasks = FacilitatorTask.objects.filter(
            facilitator=facilitator,
            created_at__date=task_date_obj
        ).select_related('actual_session', 'actual_session__planned_session', 'actual_session__planned_session__class_section').order_by('-created_at')
    else:
        # No date selected, return empty
        facilitator_tasks = FacilitatorTask.objects.none()
    
    # Get all task dates for calendar view (last 30 days)
    from django.db.models import F
    task_dates = FacilitatorTask.objects.filter(
        facilitator=facilitator,
        created_at__date__gte=start_date,
        created_at__date__lte=end_date
    ).values_list('created_at__date', flat=True).distinct()
    task_dates_list = sorted(list(set([d.day for d in task_dates if d])))
    
    # Get facilitator's feedback - FILTERED BY DATE
    facilitator_feedback = SessionFeedback.objects.filter(
        actual_session__facilitator=facilitator,
        feedback_date__gte=start_date,
        feedback_date__lte=end_date
    ).select_related('actual_session', 'actual_session__planned_session', 'actual_session__planned_session__class_section').order_by('-feedback_date')[:10]
    
    # Get lesson plan uploads - FILTERED BY DATE
    lesson_date_filter = request.GET.get('lesson_date_filter', '30')
    try:
        lesson_days = int(lesson_date_filter)
    except (ValueError, TypeError):
        lesson_days = 30
    
    lesson_end_date = timezone.now().date()
    lesson_start_date = lesson_end_date - timedelta(days=lesson_days)
    
    from .models import LessonPlanUpload
    lesson_plan_uploads = LessonPlanUpload.objects.filter(
        facilitator=facilitator,
        upload_date__gte=lesson_start_date,
        upload_date__lte=lesson_end_date
    ).select_related('planned_session', 'planned_session__class_section', 'planned_session__class_section__school').order_by('-upload_date')
    
    # Get facilitator attendance stats - FILTERED BY DATE
    from django.db.models import Count, Q
    from .models import ActualSession
    
    facilitator_sessions = ActualSession.objects.filter(
        facilitator=facilitator,
        date__gte=start_date,
        date__lte=end_date
    )
    
    total_sessions = facilitator_sessions.count()
    present_sessions = facilitator_sessions.filter(facilitator_attendance='present').count()
    absent_sessions = facilitator_sessions.filter(facilitator_attendance='absent').count()
    leave_sessions = facilitator_sessions.filter(facilitator_attendance='leave').count()
    
    # Calculate attendance rate
    facilitator_attendance_rate = 0
    if total_sessions > 0:
        facilitator_attendance_rate = round((present_sessions / total_sessions) * 100, 2)
    
    # Calculate counts for top cards
    school_count = facilitator_schools.count()
    class_count = facilitator_classes.count()
    
    # Student count logic (match facilitator_students logic but for count)
    if selected_class_id:
        student_count = Student.objects.filter(
            enrollments__class_section_id=selected_class_id,
            enrollments__is_active=True
        ).distinct().count()
    else:
        student_count = Student.objects.filter(
            enrollments__class_section__school_id__in=school_ids,
            enrollments__is_active=True
        ).distinct().count()
    
    context = {
        'facilitator': facilitator,
        'assigned_schools': facilitator_schools,
        'classes': facilitator_classes,
        'students': facilitator_students,
        'tasks': facilitator_tasks,
        'task_dates': task_dates_list,
        'feedback': facilitator_feedback,
        'lesson_plans': lesson_plan_uploads,
        'sessions': facilitator_sessions[:10],
        'selected_class': selected_class_id,
        'selected_task_date': selected_task_date,
        'date_filter': date_filter,
        'base_template': base_template, # Support dynamic sidebar
        'total_sessions': total_sessions,
        'present_sessions': present_sessions,
        'absent_sessions': absent_sessions,
        'leave_sessions': leave_sessions,
        'facilitator_attendance_rate': facilitator_attendance_rate,
        'school_count': school_count,
        'class_count': class_count,
        'student_count': student_count,
        'start_date': start_date,
        'end_date': end_date,
        # Backward compatibility for some template fields if needed
        'schools': facilitator_schools,
        'facilitator_tasks': facilitator_tasks,
        'facilitator_feedback': facilitator_feedback,
        'lesson_plan_uploads': lesson_plan_uploads,
    }
    
    return render(request, "supervisor/facilitators/detail.html", context)


@login_required
@supervisor_required
def supervisor_facilitator_detail(request, facilitator_id, base_template='supervisor/shared/base.html'):
    """Supervisor entry point for facilitator detail"""
    return _facilitator_detail_logic(request, facilitator_id, base_template=base_template)


@login_required
@admin_required
def admin_facilitator_detail(request, facilitator_id):
    """Admin entry point for facilitator detail (uses admin layout)"""
    return _facilitator_detail_logic(request, facilitator_id, base_template='admin/shared/base.html')


# Admin-accessible wrappers (Unified Navigation)
# =====================================================

@login_required
@supervisor_required
def supervisor_assign_facilitator_school(request, facilitator_id):
    """Assign facilitator to schools"""
    
    facilitator = get_object_or_404(User, id=facilitator_id, role__name__iexact="FACILITATOR")
    
    if request.method == "POST":
        school_ids = request.POST.getlist('schools')
        
        # Clear existing assignments
        FacilitatorSchool.objects.filter(facilitator=facilitator).delete()
        
        # Create new assignments
        for school_id in school_ids:
            try:
                school = School.objects.get(id=school_id)
                FacilitatorSchool.objects.create(
                    facilitator=facilitator,
                    school=school,
                    is_active=True
                )
            except School.DoesNotExist:
                pass
        
        messages.success(request, f"Facilitator '{facilitator.full_name}' assigned to schools successfully!")
        return redirect("supervisor_facilitator_detail", facilitator_id=facilitator_id)
    
    # Get all schools and currently assigned schools
    all_schools = School.objects.all().order_by('name')
    assigned_schools = FacilitatorSchool.objects.filter(
        facilitator=facilitator,
        is_active=True
    ).values_list('school_id', flat=True)
    
    context = {
        'facilitator': facilitator,
        'all_schools': all_schools,
        'assigned_schools': assigned_schools,
    }
    
    return render(request, "supervisor/facilitators/assign_schools.html", context)

@login_required
@supervisor_required
def supervisor_assign_facilitator_class(request, facilitator_id):
    """View classes in facilitator's assigned schools"""
    
    facilitator = get_object_or_404(User, id=facilitator_id, role__name__iexact="FACILITATOR")
    
    # Get facilitator's assigned schools
    assigned_schools = FacilitatorSchool.objects.filter(
        facilitator=facilitator,
        is_active=True
    ).values_list('school_id', flat=True)
    
    # Get all classes from assigned schools
    all_classes = ClassSection.objects.filter(
        school_id__in=assigned_schools
    ).select_related('school').order_by('school__name', 'class_level', 'section')
    
    context = {
        'facilitator': facilitator,
        'all_classes': all_classes,
        'assigned_schools': assigned_schools,
    }
    
    return render(request, "supervisor/facilitators/view_classes.html", context)


# =====================================================
# Class Management for Supervisor
# =====================================================
@login_required
@supervisor_required
def supervisor_class_create(request):
    """Create new class - Supervisor can add classes"""
    
    if request.method == "POST":
        form = ClassSectionForm(request.POST)
        if form.is_valid():
            class_section = form.save()
            # Invalidate cache
            cache.delete("supervisor_classes_list_all")
            cache.delete(f"supervisor_classes_list_{class_section.school_id}")
            messages.success(request, f"Class '{class_section.class_level} {class_section.section}' created successfully!")
            return redirect("supervisor_classes_list")
    else:
        form = ClassSectionForm()
    
    context = {
        'form': form,
        'title': 'Add New Class'
    }
    
    return render(request, "supervisor/classes/create.html", context)


@login_required
@supervisor_required
def supervisor_class_bulk_create(request):
    """Bulk create multiple classes at once"""
    
    schools = School.objects.all().order_by('name')
    class_levels = list(range(1, 11))  # Classes 1-10
    
    if request.method == "POST":
        school_id = request.POST.get('school')
        class_levels_selected = request.POST.getlist('class_levels')
        section = request.POST.get('section', 'A')
        academic_year = request.POST.get('academic_year', '2024-2025')
        
        if not school_id or not class_levels_selected:
            messages.error(request, "Please select a school and at least one class level")
            return redirect("supervisor_class_bulk_create")
        
        try:
            school = School.objects.get(id=school_id)
            created_count = 0
            
            for level in class_levels_selected:
                # Check if class already exists
                existing = ClassSection.objects.filter(
                    school=school,
                    class_level=level,
                    section=section,
                    academic_year=academic_year
                ).exists()
                
                if not existing:
                    ClassSection.objects.create(
                        school=school,
                        class_level=level,
                        section=section,
                        academic_year=academic_year,
                        is_active=True
                    )
                    created_count += 1
            
            if created_count > 0:
                messages.success(request, f"Successfully created {created_count} class(es) in {school.name}")
            else:
                messages.warning(request, "All selected classes already exist")
            
            return redirect("supervisor_classes_list")
        
        except School.DoesNotExist:
            messages.error(request, "School not found")
            return redirect("supervisor_class_bulk_create")
    
    context = {
        'schools': schools,
        'class_levels': class_levels,
    }
    
    return render(request, "supervisor/classes/bulk_create.html", context)


@login_required
@supervisor_required
def supervisor_class_edit(request, class_id):
    """Edit class details - Supervisor can edit classes"""
    
    class_section = get_object_or_404(ClassSection, id=class_id)
    
    if request.method == "POST":
        form = ClassSectionForm(request.POST, instance=class_section)
        if form.is_valid():
            class_section = form.save()
            # Invalidate cache
            cache.delete("supervisor_classes_list_all")
            cache.delete(f"supervisor_classes_list_{class_section.school_id}")
            messages.success(request, f"Class '{class_section.class_level} {class_section.section}' updated successfully!")
            return redirect("supervisor_classes_list")
    else:
        form = ClassSectionForm(instance=class_section)
    
    context = {
        'form': form,
        'class_section': class_section,
        'title': 'Edit Class'
    }
    
    return render(request, "supervisor/classes/edit.html", context)


@login_required
@supervisor_required
def supervisor_class_delete(request, class_id):
    """Delete class - Supervisor can delete classes"""
    
    class_section = get_object_or_404(ClassSection, id=class_id)
    school_id = class_section.school_id
    
    if request.method == "POST":
        class_name = f"{class_section.class_level} {class_section.section}"
        class_section.delete()
        # Invalidate cache
        cache.delete("supervisor_classes_list_all")
        cache.delete(f"supervisor_classes_list_{school_id}")
        messages.success(request, f"Class '{class_name}' deleted successfully!")
        return redirect("supervisor_classes_list")
    
    context = {
        'class_section': class_section,
    }
    
    return render(request, "supervisor/classes/delete_confirm.html", context)


@login_required
@supervisor_required
def supervisor_bulk_add_classes(request):
    """Bulk add multiple classes - Redirect to calendar to create grouped session"""
    
    # Get selected class IDs from query params
    selected_ids = request.GET.getlist('ids')
    
    if not selected_ids:
        messages.error(request, "No classes selected")
        return redirect("supervisor_classes_list")
    
    # Get the selected classes to verify they exist
    selected_classes = ClassSection.objects.filter(
        id__in=selected_ids
    ).select_related('school').order_by('school__name', 'class_level', 'section')
    
    if not selected_classes.exists():
        messages.error(request, "No valid classes selected")
        return redirect("supervisor_classes_list")
    
    # Redirect to calendar add date with selected class IDs as query params
    # The calendar view will handle creating the grouped session
    query_string = '&'.join([f'class_ids={id}' for id in selected_ids])
    messages.info(request, f"Creating grouped session for {len(selected_classes)} classes. Select a date and confirm.")
    return redirect(f"{reverse('supervisor_calendar_add_date')}?{query_string}")



# =====================================================
# CALENDAR MANAGEMENT
# =====================================================

def _calendar_logic(request, base_template='supervisor/shared/base.html'):
    """Core logic for supervisor calendar (undecorated)"""
    """Supervisor Calendar - View and manage calendar dates with facilitator filtering"""
    from datetime import datetime, timedelta
    from .models import SupervisorCalendar, CalendarDate
    
    # Get or create supervisor calendar
    calendar, created = SupervisorCalendar.objects.get_or_create(
        supervisor=request.user
    )
    
    # Get current month
    today = datetime.now().date()
    year = int(request.GET.get('year', today.year))
    month = int(request.GET.get('month', today.month))
    
    # Get school and facilitator filters from request
    selected_school_id = request.GET.get('school_filter', None)
    selected_facilitator_id = request.GET.get('facilitator_filter', None)
    
    # Get all schools for filter dropdown
    all_schools = School.objects.all().order_by('name')
    
    # Get facilitators for selected school
    facilitators = []
    if selected_school_id:
        facilitators = User.objects.filter(
            role__name__iexact='FACILITATOR',
            assigned_schools__school_id=selected_school_id,
            assigned_schools__is_active=True
        ).distinct().order_by('full_name')
    
    # Get all dates for this month
    from datetime import date
    first_day = date(year, month, 1)
    if month == 12:
        last_day = date(year + 1, 1, 1) - timedelta(days=1)
    else:
        last_day = date(year, month + 1, 1) - timedelta(days=1)
    
    # Filter all calendar dates for this month across all supervisors (Global View)
    calendar_dates_query = CalendarDate.objects.filter(
        date__gte=first_day,
        date__lte=last_day
    ).select_related('class_section', 'school').prefetch_related('assigned_facilitators')
    
    if selected_school_id:
        calendar_dates_query = calendar_dates_query.filter(school_id=selected_school_id)
    
    # Filter by facilitator if selected
    if selected_facilitator_id:
        calendar_dates_query = calendar_dates_query.filter(assigned_facilitators__id=selected_facilitator_id)
    
    calendar_dates = calendar_dates_query
    
    # Create a dict for easy lookup - store ALL entries for each date
    dates_dict = {}
    for cd in calendar_dates:
        if cd.date not in dates_dict:
            dates_dict[cd.date] = []
        dates_dict[cd.date].append(cd)
    
    # Build calendar grid
    calendar_grid = []
    current_date = first_day
    week = []
    
    # Add empty cells for days before month starts
    for _ in range(current_date.weekday()):
        week.append(None)
    
    while current_date <= last_day:
        week.append({
            'date': current_date,
            'calendar_date': dates_dict.get(current_date),
        })
        
        if len(week) == 7:
            calendar_grid.append(week)
            week = []
        
        current_date += timedelta(days=1)
    
    # Add remaining empty cells
    if week:
        while len(week) < 7:
            week.append(None)
        calendar_grid.append(week)
    
    # Navigation
    prev_month = month - 1 if month > 1 else 12
    prev_year = year if month > 1 else year - 1
    next_month = month + 1 if month < 12 else 1
    next_year = year if month < 12 else year + 1
    
    context = {
        'calendar': calendar,
        'calendar_grid': calendar_grid,
        'current_month': month,
        'current_year': year,
        'month_name': first_day.strftime('%B'),
        'prev_month': prev_month,
        'prev_year': prev_year,
        'next_month': next_month,
        'next_year': next_year,
        'today': today,
        'all_schools': all_schools,
        'selected_school_id': selected_school_id,
        'facilitators': facilitators,
        'selected_facilitator_id': selected_facilitator_id,
    }
    
    return render(request, "supervisor/calendar/calendar.html", context)


@login_required
@supervisor_required
def supervisor_calendar(request, base_template='supervisor/shared/base.html'):
    """Supervisor entry point for calendar"""
    return _calendar_logic(request, base_template=base_template)


def _calendar_add_date_logic(request, redirect_url_name='supervisor_calendar'):
    """Core logic for adding calendar dates (undecorated)"""


    """Add a date to calendar (session, holiday, or office work) with bulk support"""
    # Use the passed redirect_url_name for all success/error redirects
    """Add a date to calendar (session, holiday, or office work) with bulk support"""
    from .models import SupervisorCalendar, CalendarDate
    from datetime import datetime, timedelta
    
    calendar, _ = SupervisorCalendar.objects.get_or_create(
        supervisor=request.user
    )
    
    if request.method == "POST":
        # Parse dates
        date_str = request.POST.get('date')
        end_date_str = request.POST.get('end_date', '')
        time_str = request.POST.get('time', '')
        date_type = request.POST.get('date_type')
        is_bulk = request.POST.get('is_bulk') == 'on'
        
        try:
            date_obj = datetime.strptime(date_str, '%Y-%m-%d').date()
        except:
            messages.error(request, "Invalid date format")
            return redirect("supervisor_calendar")
        
        # Parse time if provided
        time_obj = None
        if time_str:
            try:
                time_obj = datetime.strptime(time_str, '%H:%M').time()
            except:
                messages.error(request, "Invalid time format")
                return redirect("supervisor_calendar")
        
        # Get list of dates to create
        dates_to_create = [date_obj]
        
        if is_bulk and end_date_str:
            try:
                end_date_obj = datetime.strptime(end_date_str, '%Y-%m-%d').date()
                if end_date_obj < date_obj:
                    messages.error(request, "End date must be after start date")
                    return redirect("supervisor_calendar")
                
                # Get selected days of week
                selected_days = request.POST.getlist('days_of_week')
                if not selected_days:
                    messages.error(request, "Please select at least one day of week for bulk creation")
                    return redirect("supervisor_calendar")
                
                selected_days = [int(d) for d in selected_days]
                
                # Generate dates for selected days
                current = date_obj
                dates_to_create = []
                while current <= end_date_obj:
                    if current.weekday() in selected_days:
                        dates_to_create.append(current)
                    current += timedelta(days=1)
                
                if not dates_to_create:
                    messages.error(request, "No dates match the selected criteria")
                    return redirect("supervisor_calendar")
            except:
                messages.error(request, "Invalid date range")
                return redirect("supervisor_calendar")
        
        # Process based on date type
        created_count = 0
        skipped_count = 0
        
        if date_type == 'session':
            school_id = request.POST.get('school')  # Gets value from hidden field
            scope = request.POST.get('scope', 'specific')  # 'all' or 'specific'
            class_ids = request.POST.getlist('class_section')
            
            if not class_ids:
                messages.error(request, "Please select at least one class")
                return redirect(redirect_url_name)
            
            try:
                school = School.objects.get(id=school_id)
            except:
                messages.error(request, "Invalid school selected")
                return redirect(redirect_url_name)
            
            # Determine which classes to create entries for
            if scope == 'all':
                classes = list(ClassSection.objects.filter(school=school, is_active=True))
            else:
                if not class_ids:
                    messages.error(request, "Please select at least one class")
                    return redirect(redirect_url_name)
                classes = list(ClassSection.objects.filter(id__in=class_ids, school=school, is_active=True))
            
            # Create a SINGLE CalendarDate entry for all grouped classes
            # Also create ONE shared 150-day PlannedSession for all grouped classes
            # Generate ONE grouped_session_id for all dates and classes
            import uuid as uuid_module
            grouped_session_id = uuid_module.uuid4()
            
            # Initialize grouped session plans ONCE for all classes
            # This creates the 150-day session plan with grouped_session_id
            if len(classes) > 1:
                # Multiple classes - create grouped session
                # Don't delete existing sessions - just add new ones for classes that don't have them
                init_result = initialize_grouped_session_plans(classes, grouped_session_id)
                if not init_result['success']:
                    messages.error(request, f"Error initializing grouped session: {init_result['error']}")
                    return redirect(redirect_url_name)
            else:
                # Single class - create individual session plan (no grouped_session_id)
                if len(classes) == 1:
                    from .session_management import SessionBulkManager
                    try:
                        # Check if sessions already exist for this class
                        existing_sessions = PlannedSession.objects.filter(class_section=classes[0]).exists()
                        
                        if not existing_sessions:
                            # Create 150 individual sessions only if they don't exist
                            sessions_to_create = []
                            for day_number in range(1, 151):
                                session = PlannedSession(
                                    class_section=classes[0],
                                    day_number=day_number,
                                    title=f"Day {day_number} Session",
                                    description=f"Session for day {day_number}",
                                    sequence_position=day_number,
                                    is_required=True,
                                    is_active=True,
                                    grouped_session_id=None  # No grouped_session_id for single class
                                )
                                sessions_to_create.append(session)
                            
                            PlannedSession.objects.bulk_create(sessions_to_create)
                    except Exception as e:
                        messages.error(request, f"Error creating session plan: {str(e)}")
                        return redirect(redirect_url_name)
            
            for date_to_create in dates_to_create:
                # Always create a new entry for each grouped session
                # Don't merge - allow multiple grouped sessions on the same day
                calendar_date = CalendarDate.objects.create(
                    calendar=calendar,
                    date=date_to_create,
                    time=time_obj,
                    date_type=DateType.SESSION,
                    school=school
                )
                # Add all selected classes to this entry
                calendar_date.class_sections.set(classes)
                created_count += 1
        
        elif date_type == 'holiday':
            holiday_name = request.POST.get('holiday_name', '').strip()
            holiday_notes = request.POST.get('holiday_notes', '').strip()
            if not holiday_name:
                messages.error(request, "Please enter holiday name")
                return redirect(redirect_url_name)
            
            # Create holiday entries (school-level, no class)
            for date_to_create in dates_to_create:
                existing = CalendarDate.objects.filter(
                    calendar=calendar,
                    date=date_to_create,
                    date_type=DateType.HOLIDAY
                ).exists()
                
                if existing:
                    skipped_count += 1
                    continue
                
                CalendarDate.objects.create(
                    calendar=calendar,
                    date=date_to_create,
                    time=time_obj,
                    date_type=DateType.HOLIDAY,
                    holiday_name=holiday_name,
                    notes=holiday_notes
                )
                created_count += 1
        
        elif date_type == 'office_work':
            task_desc = request.POST.get('office_task_description', '').strip()
            school_id = request.POST.get('school')
            facilitator_ids = request.POST.getlist('facilitators')
            
            if not task_desc:
                messages.error(request, "Please enter office task description")
                return redirect(redirect_url_name)
            
            try:
                school = School.objects.get(id=school_id) if school_id else None
            except:
                school = None
            
            # Create office work entries
            for date_to_create in dates_to_create:
                existing = CalendarDate.objects.filter(
                    calendar=calendar,
                    date=date_to_create,
                    date_type=DateType.OFFICE_WORK
                ).exists()
                
                if existing:
                    skipped_count += 1
                    continue
                
                calendar_date = CalendarDate.objects.create(
                    calendar=calendar,
                    date=date_to_create,
                    time=time_obj,
                    date_type=DateType.OFFICE_WORK,
                    office_task_description=task_desc,
                    school=school
                )
                
                # Add assigned facilitators
                if facilitator_ids:
                    facilitators = User.objects.filter(id__in=facilitator_ids, role__name__iexact="FACILITATOR")
                    calendar_date.assigned_facilitators.set(facilitators)
                
                created_count += 1
        
        # Show summary message
        if created_count > 0:
            msg = f"[OK] Successfully processed {created_count} calendar entries"
            messages.success(request, msg)
        else:
            messages.warning(request, "No calendar entries were created or updated")
        
        return redirect(redirect_url_name)
    
    # GET request - show form
    schools = School.objects.filter(status=1).order_by('name')
    classes = ClassSection.objects.filter(is_active=True).select_related('school').order_by('school__name', 'class_level', 'section')
    facilitators = User.objects.filter(role__name__iexact="FACILITATOR", is_active=True).order_by('full_name')
    
    # Check if class IDs were passed from bulk add
    pre_selected_class_ids = request.GET.getlist('class_ids')
    
    context = {
        'schools': schools,
        'classes': classes,
        'facilitators': facilitators,
        'pre_selected_class_ids': pre_selected_class_ids,
    }
    
    return render(request, "supervisor/calendar/add_date.html", context)


@login_required
@supervisor_required
def supervisor_calendar_add_date(request, redirect_url_name='supervisor_calendar'):
    """Supervisor entry point for add date"""
    return _calendar_add_date_logic(request, redirect_url_name=redirect_url_name)


def _calendar_edit_date_logic(request, date_id, redirect_url_name='supervisor_calendar'):
    """Core logic for editing calendar dates (undecorated)"""


    """Edit a calendar date entry"""
    from .models import CalendarDate, SupervisorCalendar
    
    calendar = SupervisorCalendar.objects.get(supervisor=request.user)
    calendar_date = get_object_or_404(CalendarDate, id=date_id, calendar=calendar)
    
    if request.method == "POST":
        date_type = request.POST.get('date_type')
        
        if date_type == 'session':
            class_ids = request.POST.getlist('class_section')
            if not class_ids:
                messages.error(request, "Please select at least one class")
            else:
                try:
                    classes = ClassSection.objects.filter(id__in=class_ids)
                    calendar_date.class_sections.set(classes)
                    calendar_date.date_type = DateType.SESSION
                    calendar_date.holiday_name = ''
                    calendar_date.office_task_description = ''
                    calendar_date.save()
                    messages.success(request, "Date updated successfully")
                except Exception as e:
                    messages.error(request, f"Error updating date: {str(e)}")
        
        elif date_type == 'holiday':
            holiday_name = request.POST.get('holiday_name', '').strip()
            holiday_notes = request.POST.get('holiday_notes', '').strip()
            if holiday_name:
                calendar_date.holiday_name = holiday_name
                calendar_date.notes = holiday_notes
                calendar_date.date_type = DateType.HOLIDAY
                calendar_date.class_sections.clear()
                calendar_date.office_task_description = ''
                calendar_date.save()
                messages.success(request, "Date updated successfully")
            else:
                messages.error(request, "Please enter holiday name")
        
        elif date_type == 'office_work':
            task_desc = request.POST.get('office_task_description', '').strip()
            if task_desc:
                calendar_date.office_task_description = task_desc
                calendar_date.date_type = DateType.OFFICE_WORK
                calendar_date.class_sections.clear()
                calendar_date.holiday_name = ''
                calendar_date.save()
                messages.success(request, "Date updated successfully")
            else:
                messages.error(request, "Please enter office task description")
        
        return redirect(redirect_url_name)
    
    classes = ClassSection.objects.filter(is_active=True).select_related('school').order_by('school__name', 'class_level', 'section')
    
    # Get selected class IDs for grouped sessions
    selected_class_ids = list(calendar_date.class_sections.values_list('id', flat=True)) if calendar_date.date_type == 'session' else []
    
    context = {
        'calendar_date': calendar_date,
        'classes': classes,
        'selected_class_ids': selected_class_ids,
    }
    
    return render(request, "supervisor/calendar/edit_date.html", context)


@login_required
@supervisor_required
def supervisor_calendar_edit_date(request, date_id, redirect_url_name='supervisor_calendar'):
    """Supervisor entry point for edit date"""
    return _calendar_edit_date_logic(request, date_id, redirect_url_name=redirect_url_name)


def _calendar_delete_date_logic(request, date_id, redirect_url_name='supervisor_calendar'):
    """Core logic for deleting calendar dates (undecorated)"""
    from .models import CalendarDate, SupervisorCalendar
    calendar = SupervisorCalendar.objects.get(supervisor=request.user)
    calendar_date = get_object_or_404(CalendarDate, id=date_id, calendar=calendar)
    
    calendar_date.delete()
    messages.success(request, "[OK] Calendar entry deleted successfully")
    return redirect(redirect_url_name)


@login_required
@supervisor_required
def supervisor_calendar_delete_date(request, date_id, redirect_url_name='supervisor_calendar'):
    """Supervisor entry point for delete date"""
    return _calendar_delete_date_logic(request, date_id, redirect_url_name=redirect_url_name)


@login_required
@supervisor_required
def supervisor_calendar_delete_date_confirm(request, date_id, redirect_url_name='supervisor_calendar'):
    """Delete a calendar date entry"""
    from .models import CalendarDate, SupervisorCalendar
    
    calendar = SupervisorCalendar.objects.get(supervisor=request.user)
    calendar_date = get_object_or_404(CalendarDate, id=date_id, calendar=calendar)
    
    if request.method == "POST":
        date_str = str(calendar_date.date)
        calendar_date.delete()
        messages.success(request, "[OK] Calendar entry updated successfully")
        return redirect(redirect_url_name)
    
    context = {
        'calendar_date': calendar_date,
    }
    
    return render(request, "supervisor/calendar/delete_confirm.html", context)


# =====================================================
# BULK STUDENT IMPORT FUNCTIONS
# =====================================================

@login_required
def supervisor_student_import(request, school_id):
    """Bulk import students via CSV/Excel for a specific school"""
    from django.http import HttpResponse
    import csv
    import openpyxl
    from datetime import date
    
    school = get_object_or_404(School, id=school_id)
    
    # Check if supervisor has access to this school
    if not request.user.role or request.user.role.name.upper() != "SUPERVISOR":
        messages.error(request, "Permission denied")
        return redirect("no_permission")
    
    class_sections = ClassSection.objects.filter(school=school)
    
    if request.method == "POST":
        file = request.FILES.get("file")
        if not file:
            messages.error(request, "Please upload a file")
            return redirect(request.path)
        
        ext = file.name.split(".")[-1].lower()
        
        # Parse file
        if ext == "csv":
            rows = csv.DictReader(file.read().decode("utf-8").splitlines())
        elif ext in ["xlsx", "xls"]:
            wb = openpyxl.load_workbook(file)
            sheet = wb.active
            headers = [str(cell.value).strip() for cell in sheet[1]]
            rows = []
            for r in sheet.iter_rows(min_row=2, values_only=True):
                rows.append(dict(zip(headers, r)))
        else:
            messages.error(request, "Unsupported file format. Use CSV or Excel.")
            return redirect(request.path)
        
        created_count = 0
        skipped_count = 0
        
        # Process rows
        for row in rows:
            enrollment_no = str(row.get("enrollment_number", "")).strip()
            full_name = str(row.get("full_name", "")).strip()
            gender = str(row.get("gender", "")).strip()
            class_level = str(row.get("class_level", "")).strip()
            section = str(row.get("section", "")).strip()
            start_date = row.get("start_date") or date.today()
            
            # Validate
            if not all([enrollment_no, full_name, gender, class_level, section]):
                skipped_count += 1
                continue
            
            if gender.upper() not in ["M", "F"]:
                skipped_count += 1
                continue
            
            # Find class section
            class_section = ClassSection.objects.filter(
                school=school,
                class_level=class_level,
                section=section
            ).first()
            
            if not class_section:
                skipped_count += 1
                continue
            
            # Create student
            from .models import Student, Enrollment
            student, _ = Student.objects.get_or_create(
                enrollment_number=enrollment_no,
                defaults={
                    "full_name": full_name,
                    "gender": gender.upper()
                }
            )
            
            # Create enrollment
            enrollment, created = Enrollment.objects.get_or_create(
                student=student,
                school=school,
                class_section=class_section,
                defaults={
                    "start_date": start_date,
                    "is_active": True
                }
            )
            
            if created:
                created_count += 1
        
        # Feedback
        if created_count == 0:
            messages.warning(request, f"No students imported. Skipped: {skipped_count}")
        else:
            messages.success(request, f"{created_count} students imported (Skipped: {skipped_count})")
        
        return redirect("supervisor_school_students", school_id=school_id)
    
    return render(request, "supervisor/students/import.html", {
        "school": school,
        "class_sections": class_sections
    })


@login_required
def supervisor_download_sample_csv(request):
    """Download sample CSV for student import"""
    from django.http import HttpResponse
    import csv
    
    sample_data = [
        ["enrollment_number", "full_name", "gender", "class_level", "section", "start_date"],
        ["E001", "John Doe", "M", "1", "A", "2026-01-12"],
        ["E002", "Jane Smith", "F", "1", "A", "2026-01-12"],
        ["E003", "Bob Johnson", "M", "2", "B", "2026-01-12"],
    ]
    
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="students_sample.csv"'
    
    writer = csv.writer(response)
    for row in sample_data:
        writer.writerow(row)
    
    return response



# =====================================================
# CLUSTER MANAGEMENT VIEWS
# =====================================================

@login_required
def clusters_list(request):
    """List all clusters"""
    if request.user.role.name.upper() != "SUPERVISOR":
        messages.error(request, "You do not have permission to view clusters.")
        return redirect("no_permission")
    
    clusters = Cluster.objects.all().annotate(
        school_count=Count('schools')
    ).order_by('district', 'name')
    
    return render(request, 'supervisor/clusters/list.html', {
        'clusters': clusters
    })


@login_required
def cluster_create(request):
    """Create a new cluster"""
    if request.user.role.name.upper() != "SUPERVISOR":
        messages.error(request, "You do not have permission to create clusters.")
        return redirect("no_permission")
    
    if request.method == 'POST':
        form = ClusterForm(request.POST)
        if form.is_valid():
            cluster = form.save()
            messages.success(request, f"Cluster '{cluster.name}' created successfully!")
            return redirect('clusters_list')
    else:
        form = ClusterForm()
    
    return render(request, 'supervisor/clusters/form.html', {
        'form': form,
        'title': 'Create Cluster'
    })


@login_required
def cluster_edit(request, cluster_id):
    """Edit a cluster"""
    if request.user.role.name.upper() != "SUPERVISOR":
        messages.error(request, "You do not have permission to edit clusters.")
        return redirect("no_permission")
    
    cluster = get_object_or_404(Cluster, id=cluster_id)
    
    if request.method == 'POST':
        form = ClusterForm(request.POST, instance=cluster)
        if form.is_valid():
            cluster = form.save()
            messages.success(request, f"Cluster '{cluster.name}' updated successfully!")
            return redirect('clusters_list')
    else:
        form = ClusterForm(instance=cluster)
    
    return render(request, 'supervisor/clusters/form.html', {
        'form': form,
        'cluster': cluster,
        'title': 'Edit Cluster'
    })


@login_required
def cluster_detail(request, cluster_id):
    """View cluster details and schools in it"""
    if request.user.role.name.upper() != "SUPERVISOR":
        messages.error(request, "You do not have permission to view cluster details.")
        return redirect("no_permission")
    
    cluster = get_object_or_404(Cluster, id=cluster_id)
    schools = cluster.schools.all().order_by('name')
    
    return render(request, 'supervisor/clusters/detail.html', {
        'cluster': cluster,
        'schools': schools
    })


@login_required
def cluster_delete(request, cluster_id):
    """Delete a cluster"""
    if request.user.role.name.upper() != "SUPERVISOR":
        messages.error(request, "You do not have permission to delete clusters.")
        return redirect("no_permission")
    
    cluster = get_object_or_404(Cluster, id=cluster_id)
    
    if request.method == 'POST':
        cluster_name = cluster.name
        cluster.delete()
        messages.success(request, f"Cluster '{cluster_name}' deleted successfully!")
        return redirect('clusters_list')
    
    return render(request, 'supervisor/clusters/confirm_delete.html', {
        'cluster': cluster
    })


# =====================================================
# Supervisor Sessions Management
# =====================================================
def _sessions_list_logic(request, base_template='supervisor/shared/base.html'):
    """Core logic for sessions list (undecorated) - Supports filtering by any combination of parameters."""
    from .models import ActualSession, PlannedSession, Attendance, AttendanceStatus, SessionStatus
    from django.db.models import Count, Q, Exists, OuterRef, Prefetch, F
    from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
    from django.utils import timezone
    import json
    from datetime import datetime, timedelta
    
    # Get filter parameters
    school_id = request.GET.get('school_id')
    class_id = request.GET.get('class_id')
    status_filter = request.GET.get('status', '')
    date_filter = request.GET.get('date_filter', 'all')
    page = request.GET.get('page', 1)
    
    # Build base query for PlannedSession - OPTIMIZED with minimal joins
    planned_sessions = PlannedSession.objects.filter(is_active=True)
    
    if school_id:
        planned_sessions = planned_sessions.filter(class_section__school_id=school_id)
    
    if class_id:
        planned_sessions = planned_sessions.filter(class_section_id=class_id)
    
    # Apply status filter - check ActualSession for conducted/holiday/cancelled
    if status_filter:
        if status_filter == 'conducted':
            # Sessions that have been conducted (status = 1)
            planned_sessions = planned_sessions.filter(
                actual_sessions__status=SessionStatus.CONDUCTED
            )
        elif status_filter == 'holiday':
            # Sessions marked as holiday (status = 2)
            planned_sessions = planned_sessions.filter(
                actual_sessions__status=SessionStatus.HOLIDAY
            )
        elif status_filter == 'cancelled':
            # Sessions marked as cancelled (status = 3)
            planned_sessions = planned_sessions.filter(
                actual_sessions__status=SessionStatus.CANCELLED
            )
    
    # Apply date filter - only filter if date_filter is not 'all'
    today = timezone.localdate()
    if date_filter and date_filter != 'all':
        if date_filter == 'today':
            planned_sessions = planned_sessions.filter(actual_sessions__date=today)
        elif date_filter == 'past':
            planned_sessions = planned_sessions.filter(actual_sessions__date__lt=today)
        elif date_filter == 'future':
            planned_sessions = planned_sessions.filter(actual_sessions__date__gt=today)
        elif date_filter == 'week':
            week_start = today - timedelta(days=today.weekday())
            week_end = week_start + timedelta(days=6)
            planned_sessions = planned_sessions.filter(
                actual_sessions__date__gte=week_start,
                actual_sessions__date__lte=week_end
            )
        elif date_filter == 'month':
            month_start = today.replace(day=1)
            if today.month == 12:
                month_end = month_start.replace(year=today.year + 1, month=1) - timedelta(days=1)
            else:
                month_end = month_start.replace(month=today.month + 1) - timedelta(days=1)
            planned_sessions = planned_sessions.filter(
                actual_sessions__date__gte=month_start,
                actual_sessions__date__lte=month_end
            )
    
    # Remove duplicates and order by latest date first (NULLs last), then day number
    planned_sessions = planned_sessions.distinct().select_related(
        'class_section', 'class_section__school'
    ).order_by(F('actual_sessions__date').desc(nulls_last=True), '-day_number')
        
    # Get total count before pagination
    total_count = planned_sessions.count()
    
    # Paginate: 5 sessions per page for faster initial load
    paginator = Paginator(planned_sessions, 5)
    try:
        sessions_page = paginator.page(page)
    except PageNotAnInteger:
        sessions_page = paginator.page(1)
    except EmptyPage:
        sessions_page = paginator.page(paginator.num_pages)
    
    # Prefetch related data only for current page sessions
    session_ids = [s.id for s in sessions_page.object_list]
    actual_sessions_map = {}
    if session_ids:
        # Get all actual sessions for these planned sessions
        actual_sessions = ActualSession.objects.filter(
            planned_session_id__in=session_ids
        ).select_related('facilitator').values('planned_session_id', 'status', 'date', 'facilitator__full_name')
        
        for actual in actual_sessions:
            actual_sessions_map[actual['planned_session_id']] = actual
    
    # Attach actual session data to planned sessions
    # If no ActualSession exists, set to None (will show as "Planned" in template)
    for session in sessions_page.object_list:
        session.actual_session_data = actual_sessions_map.get(session.id)
    
    # Get all schools for filter dropdown
    supervisor_schools = School.objects.all().order_by('name')
    
    # Get classes for selected school only (for cascading filter)
    selected_school_classes = []
    if school_id:
        selected_school_classes = ClassSection.objects.filter(school_id=school_id).order_by('class_level', 'section')
    
    # Build class data for JavaScript - ONLY for selected school (not all schools)
    class_data = {}
    if school_id:
        classes_with_grouping = ClassSection.objects.filter(
            school_id=school_id
        ).annotate(
            has_grouped_session=Exists(
                PlannedSession.objects.filter(
                    class_section=OuterRef('pk'),
                    grouped_session_id__isnull=False
                )
            )
        ).order_by('class_level')
        
        # Organize by school ID for JavaScript
        class_data[str(school_id)] = []
        
        for cls in classes_with_grouping:
            grouped_count = 0
            if cls.has_grouped_session:
                grouped_session_id = PlannedSession.objects.filter(
                    class_section=cls,
                    grouped_session_id__isnull=False
                ).values_list('grouped_session_id', flat=True).first()
                
                if grouped_session_id:
                    grouped_count = PlannedSession.objects.filter(
                        grouped_session_id=grouped_session_id
                    ).values('class_section').distinct().count()
            
            class_data[str(school_id)].append({
                'id': str(cls.id),
                'name': cls.display_name,
                'is_grouped': cls.has_grouped_session,
                'grouped_count': grouped_count
            })
            
    # Check if current user is an admin for dynamic UI routing
    from .decorators import _is_admin
    is_admin = _is_admin(request.user)
    
    return render(request, 'supervisor/sessions/list.html', {
        'sessions': sessions_page,
        'paginator': paginator,
        'schools': supervisor_schools,
        'selected_school_classes': selected_school_classes,
        'class_data_json': json.dumps(class_data),
        'selected_school': school_id,
        'selected_class': class_id,
        'selected_status': status_filter,
        'selected_date_filter': date_filter,
        'total_count': total_count,
        'base_template': base_template,
        'is_admin': is_admin,
    })


@login_required
@supervisor_required
def supervisor_sessions_list(request, base_template='supervisor/shared/base.html'):
    """Supervisor entry point for sessions list"""
    return _sessions_list_logic(request, base_template=base_template)


@login_required
@admin_required
def admin_sessions_list(request):
    """Admin entry point for sessions list (uses admin layout)"""
    return _sessions_list_logic(request, base_template='admin/shared/base.html')


@csrf_exempt
@login_required
@supervisor_required
def get_classes_by_school(request):
    """AJAX endpoint to get classes for a school"""
    school_id = request.GET.get('school_id')
    
    if not school_id:
        return JsonResponse({'classes': []})
    
    classes = ClassSection.objects.filter(
        school_id=school_id
    ).annotate(
        has_grouped_session=Exists(
            PlannedSession.objects.filter(
                class_section=OuterRef('pk'),
                grouped_session_id__isnull=False
            )
        )
    ).order_by('class_level').values('id', 'display_name', 'has_grouped_session')
    
    classes_list = []
    for cls in classes:
        grouped_count = 0
        if cls['has_grouped_session']:
            grouped_session_id = PlannedSession.objects.filter(
                class_section_id=cls['id'],
                grouped_session_id__isnull=False
            ).values_list('grouped_session_id', flat=True).first()
            
            if grouped_session_id:
                grouped_count = PlannedSession.objects.filter(
                    grouped_session_id=grouped_session_id
                ).values('class_section').distinct().count()
        
        classes_list.append({
            'id': str(cls['id']),
            'name': cls['display_name'],
            'is_grouped': cls['has_grouped_session'],
            'grouped_count': grouped_count
        })
    
    return JsonResponse({'classes': classes_list})


@csrf_exempt
@login_required
@supervisor_required



@login_required
@supervisor_required
def supervisor_school_sessions_analytics(request, school_id):
    """View session analytics for a school"""
    from .models import ActualSession, Attendance
    
    school = get_object_or_404(School, id=school_id)
    
    # Get all classes in this school
    classes = ClassSection.objects.filter(school=school)
    
    # Get all sessions for this school
    sessions = PlannedSession.objects.filter(
        class_section__school=school
    )
    
    # Calculate statistics
    from .models import SessionStatus, AttendanceStatus
    total_sessions = sessions.count()
    conducted_sessions = ActualSession.objects.filter(
        planned_session__class_section__school=school,
        status=SessionStatus.CONDUCTED
    ).count()
    
    holiday_sessions = ActualSession.objects.filter(
        planned_session__class_section__school=school,
        status=SessionStatus.HOLIDAY
    ).count()
    
    cancelled_sessions = ActualSession.objects.filter(
        planned_session__class_section__school=school,
        status=SessionStatus.CANCELLED
    ).count()
    
    pending_sessions = total_sessions - conducted_sessions - holiday_sessions - cancelled_sessions
    
    # Calculate average attendance
    total_attendance = Attendance.objects.filter(
        actual_session__planned_session__class_section__school=school
    ).count()
    
    present_count = Attendance.objects.filter(
        actual_session__planned_session__class_section__school=school,
        status=AttendanceStatus.PRESENT
    ).count()
    
    avg_attendance = 0
    if total_attendance > 0:
        avg_attendance = round((present_count / total_attendance) * 100, 2)
    
    # Get class-wise breakdown
    class_stats = []
    for class_section in classes:
        class_sessions = PlannedSession.objects.filter(class_section=class_section)
        class_conducted = ActualSession.objects.filter(
            planned_session__class_section=class_section,
            status=SessionStatus.CONDUCTED
        ).count()
        
        class_pending = class_sessions.count() - class_conducted
        
        class_stats.append({
            'class': class_section,
            'total_sessions': class_sessions.count(),
            'conducted': class_conducted,
            'pending': class_pending
        })
    
    return render(request, 'supervisor/sessions/school_analytics.html', {
        'school': school,
        'total_sessions': total_sessions,
        'conducted_sessions': conducted_sessions,
        'holiday_sessions': holiday_sessions,
        'cancelled_sessions': cancelled_sessions,
        'pending_sessions': pending_sessions,
        'avg_attendance': avg_attendance,
        'class_stats': class_stats
    })


@login_required
@supervisor_required
def supervisor_attendance_report_excel(request):
    """
    Generate Excel Report for Today's School Attendance (ALL SCHOOLS & CLASSES)
    Columns: School, Class, Present, Absent, Percentage, Facilitator
    """
    today = timezone.localdate()
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Attendance_{today.strftime('%Y%m%d')}"
    
    # Headers
    headers = ['School', 'Class', 'Present', 'Absent', 'Percentage', 'Facilitator']
    ws.append(headers)
    
    # Style headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    
    # Get all active classes with their schools
    classes = ClassSection.objects.filter(is_active=True).select_related('school').order_by('school__name', 'class_level', 'section')
    
    # Get today's conducted sessions
    conducted_sessions = ActualSession.objects.filter(
        date=today, 
        status=SessionStatus.CONDUCTED
    ).select_related('planned_session', 'facilitator')
    
    # Pre-fetch group mappings to correctly attribute sessions to all classes in a group
    group_ids = [s.planned_session.grouped_session_id for s in conducted_sessions if s.planned_session and s.planned_session.grouped_session_id]
    group_map = {}
    
    # Method 1: Persistent Groups (PlannedSession.grouped_session_id)
    if group_ids:
        gs_objects = GroupedSession.objects.filter(grouped_session_id__in=group_ids).prefetch_related('class_sections')
        for gs in gs_objects:
            group_map[gs.grouped_session_id] = list(gs.class_sections.values_list('id', flat=True))

    # Method 2: Dynamic Groups (CalendarDate.class_sections)
    # This is critical for groups defined by the supervisor on a specific date
    today_calendar_groups = CalendarDate.objects.filter(
        date=today, 
        class_sections__isnull=False
    ).prefetch_related('class_sections')
    
    calendar_group_map = {} # {class_id: [all_class_ids_in_same_calendar_group]}
    for cd in today_calendar_groups:
        ids = list(cd.class_sections.values_list('id', flat=True))
        if len(ids) > 1:
            for cid in ids:
                calendar_group_map[cid] = ids

    # Build mapping: {class_id: session_object}
    # Step 1: Map classes to their direct actual sessions
    today_sessions = {
        session.planned_session.class_section_id: session 
        for session in conducted_sessions 
        if session.planned_session
    }
    
    # Step 2: Global Group Lookup
    # Map all classes in a group to ANY conducted session found for that group today.
    for session in conducted_sessions:
        if not session.planned_session:
            continue
            
        # Check Persistent Grouping
        gid = session.planned_session.grouped_session_id
        if gid and gid in group_map:
            for cls_id in group_map[gid]:
                if cls_id not in today_sessions:
                    today_sessions[cls_id] = session
        
        # Check Dynamic Calendar Grouping
        own_cid = session.planned_session.class_section_id
        if own_cid in calendar_group_map:
            for cls_id in calendar_group_map[own_cid]:
                if cls_id not in today_sessions:
                    today_sessions[cls_id] = session
    
    # Get facilitator mappings for schools (for classes without sessions)
    # {school_id: [facilitator_names]}
    school_facilitators = {}
    fac_ships = FacilitatorSchool.objects.filter(is_active=True).select_related('facilitator', 'school')
    for fs in fac_ships:
        if fs.school_id not in school_facilitators:
            school_facilitators[fs.school_id] = []
        school_facilitators[fs.school_id].append(fs.facilitator.full_name)

    for class_section in classes:
        school = class_section.school
        session = today_sessions.get(class_section.id)
        
        total_students = Enrollment.objects.filter(class_section=class_section, is_active=True).count()
        
        if session:
            # Data from conducted session - Filtered by this specific class
            # Revert to enrollment__class_section_id as denormalized field might be NULL due to bulk operations
            present_count = Attendance.objects.filter(
                actual_session=session, 
                status=AttendanceStatus.PRESENT,
                enrollment__class_section_id=class_section.id # Follow FK to be safe
            ).count()
            facilitator_name = session.facilitator.full_name if session.facilitator else "N/A"
        else:
            # No session conducted today
            present_count = 0
            # Get facilitators assigned to this school as a fallback
            facilitator_list = school_facilitators.get(school.id, [])
            facilitator_name = ", ".join(facilitator_list) if facilitator_list else "Not Assigned"
            
        absent_count = total_students - present_count
        percentage = (present_count / total_students * 100) if total_students > 0 else 0
        
        ws.append([
            school.name,
            f"{class_section.class_level} - {class_section.section}",
            present_count,
            absent_count,
            f"{percentage:.1f}%",
            facilitator_name
        ])
    
    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column].width = max_length + 2
    
    # Prepare response
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="School_Attendance_Full_{today.strftime("%Y-%m-%d")}.xlsx"'
    
    return response


@login_required
@supervisor_required
def supervisor_get_notifications(request):
    """
    API endpoint to get class unavailable notifications for supervisor
    Returns notifications from both cache and database
    """
    from django.core.cache import cache
    
    try:
        # Get notifications from cache (immediate/ephemeral)
        cache_key = f"supervisor_notifications_{request.user.id}"
        cache_notifications = cache.get(cache_key, [])
        
        # Get notifications from database (persistent/reliable across workers)
        from django.utils import timezone
        from .models import ActualSession, SessionStatus
        today = timezone.localdate()
        
        db_sessions = ActualSession.objects.filter(
            date=today,
            status=SessionStatus.CANCELLED
        ).select_related('planned_session__class_section__school', 'facilitator').order_by('-created_at')
        
        db_notifications = []
        for s in db_sessions:
            if not s.planned_session: continue
            
            db_notifications.append({
                'id': str(s.id),
                'type': 'class_unavailable',
                'class_name': s.planned_session.class_section.display_name,
                'facilitator_name': s.facilitator.full_name if s.facilitator else 'Unknown',
                'school_name': s.planned_session.class_section.school.name,
                'timestamp': s.created_at.isoformat(),
                'message': s.remarks.replace('NOTIFICATION: ', '') if s.remarks else f"Class not available reported"
            })
            
        # Merge - DB notifications are preferred for reliability
        # Using a dictionary to avoid duplicates between cache and DB
        merged = { n['id']: n for n in cache_notifications }
        for n in db_notifications:
            merged[n['id']] = n
            
        final_list = sorted(merged.values(), key=lambda x: x['timestamp'], reverse=True)
        
        return JsonResponse({
            'success': True,
            'notifications': final_list[:50],
            'count': len(final_list)
        })
    except Exception as e:
        logger.error(f"Error in supervisor_get_notifications: {e}", exc_info=True)
        return JsonResponse({
            'success': False,
            'error': str(e),
            'notifications': []
        }, status=500)


# =====================================================
# Admin-Accessible Wrappers for Supervisor Views
# These sit under /admin/* so the URLAccessControlMiddleware allows them.
# They use @admin_required and call the supervisor business logic directly.
# =====================================================

@login_required
@admin_required
def admin_view_facilitators(request):
    """Admin wrapper: facilitator list with admin sidebar"""
    return _facilitators_list_logic(request, base_template='admin/shared/base.html')


@login_required
@admin_required
def admin_view_calendar(request):
    """Admin wrapper: renders the supervisor calendar with the admin sidebar."""
    return _calendar_logic(request, base_template='admin/shared/base.html')


@login_required
@admin_required
def admin_calendar_add_date(request):
    """Admin wrapper: add calendar date with admin redirect"""
    return _calendar_add_date_logic(request, redirect_url_name='admin_view_calendar')


@login_required
@admin_required
def admin_calendar_edit_date(request, date_id):
    """Admin wrapper: edit calendar date with admin redirect"""
    # Note: supervisor_calendar_edit_date needs the same redirect_url_name treatment
    return _calendar_edit_date_logic(request, date_id, redirect_url_name='admin_view_calendar')


@login_required
@admin_required
def admin_calendar_delete_date(request, date_id):
    """Admin wrapper: delete calendar date with admin redirect"""
    # Note: supervisor_calendar_delete_date needs the same redirect_url_name treatment
    return _calendar_delete_date_logic(request, date_id, redirect_url_name='admin_view_calendar')


@login_required
@admin_required
def admin_session_detail(request, session_id):
    """Admin wrapper: session detail with admin sidebar"""
    return _session_detail_logic(request, session_id, base_template='admin/shared/base.html')


@login_required
@admin_required
def admin_get_notifications(request):
    """
    Admin wrapper: returns supervisor-style notifications JSON for admin.
    Now includes both cache and database-backed notifications for parity with supervisor view.
    """
    from django.core.cache import cache
    
    try:
        # Get notifications from cache
        cache_key = f"supervisor_notifications_{request.user.id}"
        cache_notifications = cache.get(cache_key, [])
        
        # Get notifications from database for today
        from django.utils import timezone
        from .models import ActualSession, SessionStatus
        today = timezone.localdate()
        
        db_sessions = ActualSession.objects.filter(
            date=today,
            status=SessionStatus.CANCELLED
        ).select_related('planned_session__class_section__school', 'facilitator').order_by('-created_at')
        
        db_notifications = []
        for s in db_sessions:
            if not s.planned_session: continue
            
            db_notifications.append({
                'id': str(s.id),
                'type': 'class_unavailable',
                'class_name': s.planned_session.class_section.display_name,
                'facilitator_name': s.facilitator.full_name if s.facilitator else 'Unknown',
                'school_name': s.planned_session.class_section.school.name,
                'timestamp': s.created_at.isoformat(),
                'message': s.remarks.replace('NOTIFICATION: ', '') if s.remarks else f"Class not available reported"
            })
            
        # Merge
        merged = { n['id']: n for n in cache_notifications }
        for n in db_notifications:
            merged[n['id']] = n
            
        final_list = sorted(merged.values(), key=lambda x: x.get('timestamp', ''), reverse=True)
        
        return JsonResponse({
            'success': True,
            'notifications': final_list[:50],
            'count': len(final_list)
        })
    except Exception as e:
        logger.error(f"Error in admin_get_notifications: {e}", exc_info=True)
        return JsonResponse({'success': True, 'notifications': [], 'count': 0})



@login_required
@admin_required
def admin_sessions_list(request):
    """Admin wrapper: supervisor sessions list with admin sidebar"""
    return _sessions_list_logic(request, base_template='admin/shared/base.html')


@login_required
@admin_required
def admin_class_sessions(request, class_id):
    """Admin wrapper: class sessions with admin sidebar"""
    return _class_sessions_logic(request, class_id, base_template='admin/shared/base.html')
